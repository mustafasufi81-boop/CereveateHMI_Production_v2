"""Report service for Daily Report generation and Excel export."""

from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import os


class ReportService:
    """Build Daily report data and export it as XLSX."""

    def __init__(self, connection, app_config: Dict[str, Any] | None = None):
        self.connection = connection
        self.app_config = app_config or {}

    @staticmethod
    def _hour_columns() -> List[str]:
        """Return hour columns starting from 5 AM to 4 AM (24 hours)"""
        return [
            "5 am To 6 am",
            "6 am To 7 am",
            "7 am To 8 am",
            "8 am To 9 am",
            "9 am To 10 am",
            "10 am To 11 am",
            "11 am To 12 pm",
            "12 pm To 1 pm",
            "1 pm To 2 pm",
            "2 pm To 3 pm",
            "3 pm To 4 pm",
            "4 pm To 5 pm",
            "5 pm To 6 pm",
            "6 pm To 7 pm",
            "7 pm To 8 pm",
            "8 pm To 9 pm",
            "9 pm To 10 pm",
            "10 pm To 11 pm",
            "11 pm To 12 am",
            "12 am To 1 am",
            "1 am To 2 am",
            "2 am To 3 am",
            "3 am To 4 am",
            "4 am To 5 am",
        ]

    @staticmethod
    def _ordered_hours() -> List[int]:
        """Return hour order starting from 5 AM (hour 5) to 4 AM (hour 4)"""
        return [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 0, 1, 2, 3, 4]

    def build_daily_report(
        self,
        report_date: datetime.date,
        plant: str | list[str],
        area: str | list[str],
        generated_by: str,
        page: int | None = None,
        page_size: int | None = None,
        source_id: str | None = None,
    ) -> Dict[str, Any]:
        if not self.connection:
            raise RuntimeError("Database not connected")

        # Normalize plant and area to lists for multi-select support
        plants = [plant] if isinstance(plant, str) else plant
        areas = [area] if isinstance(area, str) else area
        
        next_date = report_date + timedelta(days=1)

        with self.connection.cursor() as cursor:
            template_rows = []

            if not source_id:
                # Build IN clause for multiple areas and plants
                area_placeholders = ",".join(["%s"] * len(areas))
                plant_placeholders = ",".join(["%s"] * len(plants))
                cursor.execute(
                    f"""
                    SELECT
                        vt.s_no,
                        vt.tag_id,
                        vt.display_label,
                        vt.group_name,
                        vt.parameter_unit,
                        vt.plant,
                        vt.area,
                        tm.sub_equipment,
                        tm.description,
                        tm.eng_unit
                    FROM historian_meta.v_report_template_tags vt
                    LEFT JOIN historian_meta.tag_master tm ON vt.tag_id = tm.tag_id
                    WHERE vt.report_type = 'DAILY'
                      AND vt.plant IN ({plant_placeholders})
                      AND vt.area IN ({area_placeholders})
                      AND vt.template_enabled = TRUE
                      AND vt.tag_enabled = TRUE
                      AND vt.tag_id IN (
                          SELECT DISTINCT tag_id 
                          FROM historian_raw.v_daily_hourly_agg
                          WHERE local_date = %s
                      )
                    ORDER BY vt.group_name ASC, vt.s_no ASC
                    """,
                    (*plants, *areas, report_date),
                )
                template_rows = cursor.fetchall()

            if not template_rows:
                source_filter = "AND tm.server_progid = %s" if source_id else ""
                area_placeholders = ",".join(["%s"] * len(areas))
                plant_placeholders = ",".join(["%s"] * len(plants))
                fallback_params = (*plants, *areas, source_id) if source_id else (*plants, *areas)
                cursor.execute(
                    f"""
                    SELECT
                        ROW_NUMBER() OVER (ORDER BY tm.equipment, tm.tag_name) AS s_no,
                        tm.tag_id,
                        tm.tag_name AS display_label,
                        tm.equipment AS group_name,
                        COALESCE(tm.eng_unit, tm.data_type, '') AS parameter_unit,
                        tm.plant,
                        tm.area,
                        tm.sub_equipment,
                        tm.description,
                        tm.eng_unit
                    FROM historian_meta.tag_master tm
                    WHERE tm.plant IN ({plant_placeholders})
                      AND tm.area IN ({area_placeholders})
                      {source_filter}
                      AND tm.enabled = TRUE
                      AND COALESCE(tm.include_in_report, TRUE) = TRUE
                      AND tm.tag_id IN (
                          SELECT DISTINCT tag_id 
                          FROM historian_raw.v_daily_hourly_agg
                          WHERE local_date = %s
                      )
                    ORDER BY tm.equipment, tm.tag_name
                    """,
                    fallback_params + (report_date,) if source_id else fallback_params + (report_date,),
                )
                template_rows = cursor.fetchall()

            if not template_rows:
                return {
                    "meta": {
                        "company": self.app_config.get("reporting", {}).get("company_name", "BHARAT ALUMINIUM COMPANY LIMITED"),
                        "plant": ', '.join(plants),
                        "report_title": f"Utility Report Of :- {', '.join(areas)}",
                        "date": report_date.isoformat(),
                        "generated_at": datetime.now().isoformat(),
                        "generated_by": generated_by,
                    },
                    "columns": self._hour_columns(),
                    "rows": [],
                    "pagination": {
                        "page": 1,
                        "page_size": int(page_size) if page_size else 0,
                        "total_rows": 0,
                        "total_pages": 0,
                    },
                }

            tag_ids = [row["tag_id"] for row in template_rows]

            cursor.execute(
                """
                SELECT
                    tag_id,
                    local_date,
                    local_hour AS hour,
                    avg_val,
                    max_val,
                    min_val
                FROM historian_raw.v_daily_hourly_agg
                WHERE tag_id = ANY(%s)
                  AND local_date = %s
                ORDER BY tag_id, local_date, local_hour
                """,
                (tag_ids, report_date),
            )
            agg_rows = cursor.fetchall()
            
            # DEBUG: Log data retrieval
            print(f"[DEBUG] build_daily_report: date={report_date}, plants={plants}, areas={areas}")
            print(f"[DEBUG] Template tags found: {len(template_rows)}")
            print(f"[DEBUG] Aggregated rows from view: {len(agg_rows)}")
            if agg_rows:
                unique_tags = set(row["tag_id"] for row in agg_rows)
                print(f"[DEBUG] Unique tags with data: {len(unique_tags)}")

        hourly_map: Dict[str, Dict[int, Dict[str, Any]]] = {}
        for row in agg_rows:
            tag_id = row["tag_id"]
            hour = int(row["hour"])
            hourly_map.setdefault(tag_id, {})[hour] = {
                "avg": float(row["avg_val"]) if row["avg_val"] is not None else None,
                "max": float(row["max_val"]) if row["max_val"] is not None else None,
                "min": float(row["min_val"]) if row["min_val"] is not None else None,
            }

        rows_out: List[Dict[str, Any]] = []
        for trow in template_rows:
            tag_id = trow["tag_id"]
            hour_data = hourly_map.get(tag_id, {})
            ordered_hours = self._ordered_hours()
            hourly_values: List[float | None] = []
            all_max: List[float] = []
            all_min: List[float] = []
            all_avg: List[float] = []

            for h in ordered_hours:
                item = hour_data.get(h)
                if item and item["avg"] is not None:
                    hourly_values.append(item["avg"])
                    all_avg.append(item["avg"])
                    if item["max"] is not None:
                        all_max.append(item["max"])
                    if item["min"] is not None:
                        all_min.append(item["min"])
                else:
                    hourly_values.append(None)

            row_avg = round(sum(all_avg) / len(all_avg), 2) if all_avg else None
            row_max = round(max(all_max), 2) if all_max else None
            row_min = round(min(all_min), 2) if all_min else None

            # Helper to filter out 'None' string and None/empty values
            def clean_field(value):
                if value is None or value == '' or value == 'None':
                    return ''
                return value
            
            rows_out.append(
                {
                    "s_no": trow["s_no"],
                    "group": clean_field(trow.get("group_name")),
                    "sub_equipment": clean_field(trow.get("sub_equipment")),
                    "tag_id": tag_id,
                    "description": clean_field(trow.get("description")) or clean_field(trow.get("display_label")) or tag_id,
                    "eng_unit": clean_field(trow.get("eng_unit")),
                    "parameter_unit": clean_field(trow.get("parameter_unit")),
                    "display_label": clean_field(trow.get("display_label")) or tag_id,
                    "avg": row_avg,
                    "max": row_max,
                    "min": row_min,
                    "hourly": hourly_values,
                }
            )

        meta = {
            "company": self.app_config.get("reporting", {}).get("company_name", "BHARAT ALUMINIUM COMPANY LIMITED"),
            "plant": ', '.join(plants),
            "report_title": f"Utility Report Of :- {', '.join(areas)}",
            "date": report_date.isoformat(),
            "generated_at": datetime.now().isoformat(),
            "generated_by": generated_by,
        }

        total_rows = len(rows_out)
        rows_for_response = rows_out
        pagination = {
            "page": 1,
            "page_size": total_rows,
            "total_rows": total_rows,
            "total_pages": 1 if total_rows else 0,
        }

        if page is not None and page_size is not None:
            safe_page_size = max(1, int(page_size))
            total_pages = max(1, (total_rows + safe_page_size - 1) // safe_page_size)
            safe_page = min(max(1, int(page)), total_pages)
            start = (safe_page - 1) * safe_page_size
            end = start + safe_page_size
            rows_for_response = rows_out[start:end]
            pagination = {
                "page": safe_page,
                "page_size": safe_page_size,
                "total_rows": total_rows,
                "total_pages": total_pages,
            }

        return {
            "meta": meta,
            "columns": self._hour_columns(),
            "rows": rows_for_response,
            "pagination": pagination,
        }

    def export_to_excel(self, report_data: Dict[str, Any], tag_ids: list | None = None) -> BytesIO:
        # Apply tag filter if provided (download-filtered feature)
        if tag_ids:
            report_data = dict(report_data)
            report_data["rows"] = [r for r in report_data["rows"] if r["tag_id"] in tag_ids]

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Add company logo on the left side (Row 1)
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "apex-hmi", "public", "Logo_Company.png")
        if os.path.exists(logo_path):
            img = Image(logo_path)
            # Scale logo to fit nicely (height ~60 pixels)
            img.height = 60
            img.width = int(img.width * (60 / img.height))
            ws.add_image(img, "A1")
        
        # Row 1: Company name and plant (centered, leaving space for logo)
        ws.append(["", "", "BHARAT ALUMINIUM COMPANY LIMITED ( PLANT- II )", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws.merge_cells("C1:N1")
        ws["C1"].font = Font(bold=True, size=14)
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Row 2: Plant section and date (centered)
        ws.append(["", "", "POTLINE, FUME TREATMENT PLANT", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws.merge_cells("C2:N2")
        ws["C2"].font = Font(bold=True, size=12)
        ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        # Row 3: Report title and date (centered)
        date_obj = datetime.strptime(report_data["meta"]["date"], "%Y-%m-%d")
        date_str = date_obj.strftime("%d-%B-%y")
        ws.append(["", "", "DAILY REPORT  (CONTROL ROOM)", "", "", "", "", "", "", "", "", "", f"DATE: {date_str}", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws.merge_cells("C3:K3")
        ws["C3"].font = Font(bold=True, size=12)
        ws["C3"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("M3:P3")
        ws["M3"].font = Font(bold=True, size=11)
        ws["M3"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 20

        # Row 4: Column headers with proper formatting
        # EXACT FORMAT: Equipment | Sub Equipment | Tag Name | Tag Description | Unit | 6 AM to 7 AM | ... | MIN | MAX | AVG
        hour_labels = []
        for hour_col in report_data["columns"]:
            parts = hour_col.split(" To ")
            if len(parts) == 2:
                hour_labels.append(f"{parts[0].upper()} \n To {parts[1].upper()}")
            else:
                hour_labels.append(hour_col)
        
        # COLUMN ORDER: Equipment | Sub Equipment | Tag Name | Tag Description | Unit | AVG | MIN | MAX | [hourly]
        headers_row = ["Equipment", "Sub Equipment", "Tag Name", "Tag Description", "Unit", "AVG", "MIN", "MAX"] + hour_labels
        ws.append(headers_row)
        
        # Style header row
        header_fill = PatternFill("solid", fgColor="4472C4")
        summary_fill = PatternFill("solid", fgColor="E7E6E6")  # Light gray for AVG/MIN/MAX columns
        white_bold = Font(color="FFFFFF", bold=True, size=10)
        for cell in ws[4]:
            if cell.value:
                cell.fill = header_fill
                cell.font = white_bold
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Helper to filter out 'None' string and None/empty values
        def clean_field(value):
            if value is None or value == '' or value == 'None':
                return ''
            return value
        
        # Data rows - Format: Equipment | Sub Equipment | Tag Name | Tag Description | Unit | AVG | MIN | MAX | [hourly]
        sorted_rows = sorted(
            report_data["rows"],
            key=lambda r: (
                (r.get("group") or "").lower(),
                (r.get("display_label") or r.get("tag_id") or "").lower(),
                (r.get("parameter_unit") or "").lower(),
            ),
        )

        for row in sorted_rows:
            data_row = [
                clean_field(row.get("group")),                      # Equipment
                clean_field(row.get("sub_equipment")),              # Sub Equipment
                row["tag_id"],                                      # Tag Name
                clean_field(row.get("description")) or clean_field(row.get("display_label")) or row["tag_id"],  # Tag Description
                clean_field(row.get("eng_unit")),                   # Unit
                row["avg"],                                         # AVG  ← col 6
                row["min"],                                         # MIN  ← col 7
                row["max"],                                         # MAX  ← col 8
                *row["hourly"],                                     # Hourly values ← col 9+
            ]
            ws.append(data_row)

        start_data_row = 5
        end_data_row = ws.max_row
        
        # Apply borders and number formatting to all data cells
        # Columns 1-5: fixed meta, 6=AVG, 7=MIN, 8=MAX, 9+=hourly
        for r in range(start_data_row, end_data_row + 1):
            for c in range(1, len(headers_row) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                # Number formatting for AVG/MIN/MAX and hourly columns
                if c >= 6 and isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                # Highlight AVG, MIN, MAX columns (fixed positions 6, 7, 8)
                if c in (6, 7, 8):
                    cell.fill = summary_fill
                # Center align Equipment, Sub Equipment, and Unit columns
                if c in (1, 2, 5):
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # Left align Tag Name and Tag Description
                if c in (3, 4):
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Set column widths (updated for new structure)
        widths = {
            "A": 22,  # Equipment
            "B": 18,  # Sub Unit
            "C": 24,  # Tag Name
            "D": 40,  # Tag Description
            "E": 12,  # Unit
            "F": 14,  # First hour column
            "G": 14,  # Second hour column
        }
        # All hourly columns and stats
        for col in "HIJKLMNOPQRSTUVWXYZ":
            widths[col] = 14
        widths["AA"] = 14  # MIN
        widths["AB"] = 14  # MAX
        widths["AC"] = 14  # AVG
        widths["AD"] = 14
        widths["AE"] = 14

        for key, value in widths.items():
            ws.column_dimensions[key].width = value

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    # ------------------------------------------------------------------ #
    # Monthly Report                                                       #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _date_columns(from_date: datetime.date, to_date: datetime.date) -> List[str]:
        days = (to_date - from_date).days + 1
        return [(from_date + timedelta(days=idx)).isoformat() for idx in range(days)]

    def build_monthly_report(
        self,
        from_date: datetime.date,
        to_date: datetime.date,
        plant: str | list[str],
        area: str | list[str],
        generated_by: str,
        page: int | None = None,
        page_size: int | None = None,
        source_id: str | None = None,
    ) -> Dict[str, Any]:
        if not self.connection:
            raise RuntimeError("Database not connected")

        # Normalize plant and area to lists for multi-select support
        plants = [plant] if isinstance(plant, str) else plant
        areas = [area] if isinstance(area, str) else area

        columns = self._date_columns(from_date, to_date)
        next_to_date = to_date + timedelta(days=1)

        with self.connection.cursor() as cursor:
            template_rows = []

            if not source_id:
                # Build IN clause for multiple areas and plants
                area_placeholders = ",".join(["%s"] * len(areas))
                plant_placeholders = ",".join(["%s"] * len(plants))
                cursor.execute(
                    f"""
                    SELECT
                        vt.s_no,
                        vt.tag_id,
                        vt.display_label,
                        vt.group_name,
                        vt.parameter_unit,
                        vt.plant,
                        vt.area,
                        tm.sub_equipment,
                        tm.description,
                        tm.eng_unit
                    FROM historian_meta.v_report_template_tags vt
                    LEFT JOIN historian_meta.tag_master tm ON vt.tag_id = tm.tag_id
                    WHERE vt.report_type = 'MONTHLY'
                      AND vt.plant IN ({plant_placeholders})
                      AND vt.area IN ({area_placeholders})
                      AND vt.template_enabled = TRUE
                      AND vt.tag_enabled = TRUE
                    ORDER BY vt.group_name ASC, vt.s_no ASC
                    """,
                    (*plants, *areas),
                )
                template_rows = cursor.fetchall()

            if not template_rows and not source_id:
                area_placeholders = ",".join(["%s"] * len(areas))
                plant_placeholders = ",".join(["%s"] * len(plants))
                cursor.execute(
                    f"""
                    SELECT
                        vt.s_no,
                        vt.tag_id,
                        vt.display_label,
                        vt.group_name,
                        vt.parameter_unit,
                        vt.plant,
                        vt.area,
                        tm.sub_equipment,
                        tm.description,
                        tm.eng_unit
                    FROM historian_meta.v_report_template_tags vt
                    LEFT JOIN historian_meta.tag_master tm ON vt.tag_id = tm.tag_id
                    WHERE vt.report_type = 'DAILY'
                      AND vt.plant IN ({plant_placeholders})
                      AND vt.area IN ({area_placeholders})
                      AND vt.template_enabled = TRUE
                      AND vt.tag_enabled = TRUE
                    ORDER BY vt.group_name ASC, vt.s_no ASC
                    """,
                    (*plants, *areas),
                )
                template_rows = cursor.fetchall()

            if not template_rows:
                # Fallback: use ALL enabled tags from tag_master for this plant/area that have data
                # Optionally filtered by server_progid (source/topic)
                source_filter = "AND tm.server_progid = %s" if source_id else ""
                area_placeholders = ",".join(["%s"] * len(areas))
                plant_placeholders = ",".join(["%s"] * len(plants))
                fallback_params = (*plants, *areas, source_id) if source_id else (*plants, *areas)
                cursor.execute(
                    f"""
                    SELECT
                        ROW_NUMBER() OVER (ORDER BY tm.equipment, tm.tag_name) AS s_no,
                        tm.tag_id,
                        tm.tag_name AS display_label,
                        tm.equipment AS group_name,
                        COALESCE(tm.eng_unit, tm.data_type, '') AS parameter_unit,
                        tm.plant,
                        tm.area,
                        tm.sub_equipment,
                        tm.description,
                        tm.eng_unit
                    FROM historian_meta.tag_master tm
                    WHERE tm.plant IN ({plant_placeholders})
                      AND tm.area IN ({area_placeholders})
                      {source_filter}
                      AND tm.enabled = TRUE
                      AND COALESCE(tm.include_in_report, TRUE) = TRUE
                      AND tm.tag_id IN (
                          SELECT DISTINCT tag_id FROM historian_raw.v_daily_hourly_agg
                      )
                    ORDER BY tm.equipment, tm.tag_name
                    """,
                    fallback_params,
                )
                template_rows = cursor.fetchall()

            if not template_rows:
                return {
                    "meta": {
                        "company": self.app_config.get("reporting", {}).get("company_name", "BHARAT ALUMINIUM COMPANY LIMITED"),
                        "plant": ', '.join(plants),
                        "report_title": f"Utility Report Of :- {', '.join(areas)}",
                        "from_date": from_date.isoformat(),
                        "to_date": to_date.isoformat(),
                        "generated_at": datetime.now().isoformat(),
                        "generated_by": generated_by,
                    },
                    "columns": columns,
                    "rows": [],
                    "pagination": {
                        "page": 1,
                        "page_size": int(page_size) if page_size else 0,
                        "total_rows": 0,
                        "total_pages": 0,
                    },
                }

            tag_ids = [row["tag_id"] for row in template_rows]

            cursor.execute(
                """
                SELECT
                    tag_id,
                    local_date,
                    ROUND(AVG(avg_val)::NUMERIC, 2) AS daily_avg,
                    ROUND(MAX(max_val)::NUMERIC, 2) AS daily_max,
                    ROUND(MIN(min_val)::NUMERIC, 2) AS daily_min
                FROM historian_raw.v_daily_hourly_agg
                WHERE tag_id = ANY(%s)
                  AND local_date >= %s
                  AND local_date <= %s
                GROUP BY tag_id, local_date
                ORDER BY tag_id, local_date
                """,
                (tag_ids, from_date, to_date),
            )
            agg_rows = cursor.fetchall()

        # Map tag -> date -> daily aggregated values (avg, max, min)
        daily_inputs: Dict[str, Dict[datetime.date, Dict[str, Any]]] = {}
        for row in agg_rows:
            tag_id = row["tag_id"]
            local_date = row["local_date"]
            daily_avg = float(row["daily_avg"]) if row["daily_avg"] is not None else None
            daily_max = float(row["daily_max"]) if row["daily_max"] is not None else None
            daily_min = float(row["daily_min"]) if row["daily_min"] is not None else None
            if daily_avg is None:
                continue
            daily_inputs.setdefault(tag_id, {})[local_date] = {
                "avg": daily_avg,
                "max": daily_max,
                "min": daily_min,
            }

        # Helper to filter out 'None' string and None/empty values
        def clean_field(value):
            if value is None or value == '' or value == 'None':
                return ''
            return value

        rows_out: List[Dict[str, Any]] = []
        for trow in template_rows:
            tag_id = trow["tag_id"]
            by_day = daily_inputs.get(tag_id, {})

            day_values: List[float | None] = []
            all_avg: List[float] = []
            all_max: List[float] = []
            all_min: List[float] = []
            for col in columns:
                report_day = datetime.strptime(col, "%Y-%m-%d").date()
                stats = by_day.get(report_day)
                if stats is not None:
                    day_values.append(stats["avg"])
                    all_avg.append(stats["avg"])
                    if stats["max"] is not None:
                        all_max.append(stats["max"])
                    if stats["min"] is not None:
                        all_min.append(stats["min"])
                else:
                    day_values.append(None)

            rows_out.append(
                {
                    "s_no": trow["s_no"],
                    "group": clean_field(trow.get("group_name")),
                    "sub_equipment": clean_field(trow.get("sub_equipment")),
                    "tag_id": tag_id,
                    "description": clean_field(trow.get("description")) or clean_field(trow.get("display_label")) or tag_id,
                    "eng_unit": clean_field(trow.get("eng_unit")),
                    "parameter_unit": clean_field(trow.get("parameter_unit")),
                    "display_label": clean_field(trow.get("display_label")) or tag_id,
                    "avg": round(sum(all_avg) / len(all_avg), 2) if all_avg else None,
                    "max": round(max(all_max), 2) if all_max else None,
                    "min": round(min(all_min), 2) if all_min else None,
                    "hourly": day_values,
                }
            )

        meta = {
            "company": self.app_config.get("reporting", {}).get("company_name", "BHARAT ALUMINIUM COMPANY LIMITED"),
            "plant": ', '.join(plants),
            "report_title": f"Utility Report Of :- {', '.join(areas)}",
            "from_date": from_date.isoformat(),
            "to_date": to_date.isoformat(),
            "generated_at": datetime.now().isoformat(),
            "generated_by": generated_by,
        }

        total_rows = len(rows_out)
        rows_for_response = rows_out
        pagination = {
            "page": 1,
            "page_size": total_rows,
            "total_rows": total_rows,
            "total_pages": 1 if total_rows else 0,
        }

        if page is not None and page_size is not None:
            safe_page_size = max(1, int(page_size))
            total_pages = max(1, (total_rows + safe_page_size - 1) // safe_page_size)
            safe_page = min(max(1, int(page)), total_pages)
            start = (safe_page - 1) * safe_page_size
            rows_for_response = rows_out[start: start + safe_page_size]
            pagination = {
                "page": safe_page,
                "page_size": safe_page_size,
                "total_rows": total_rows,
                "total_pages": total_pages,
            }

        return {
            "meta": meta,
            "columns": columns,
            "rows": rows_for_response,
            "pagination": pagination,
        }

    def export_monthly_to_excel(self, report_data: Dict[str, Any], tag_ids: list | None = None) -> BytesIO:
        # Apply tag filter if provided (download-filtered feature)
        if tag_ids:
            report_data = dict(report_data)
            report_data["rows"] = [r for r in report_data["rows"] if r["tag_id"] in tag_ids]

        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Report"

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Add company logo on the left side (Row 1)
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "apex-hmi", "public", "Logo_Company.png")
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.height = 60
            img.width = int(img.width * (60 / img.height))
            ws.add_image(img, "A1")
        
        # Row 1: Company name and plant (centered, leaving space for logo)
        total_cols = 5 + len(report_data["columns"]) + 3  # Equipment, Sub Equipment, Tag Name, Description, Unit + days + Min,Max,Avg
        ws.append(["", "", "BHARAT ALUMINIUM COMPANY LIMITED ( PLANT- II )"] + [""] * (total_cols - 3))
        merge_end = get_column_letter(min(14, total_cols))
        ws.merge_cells(f"C1:{merge_end}1")
        ws["C1"].font = Font(bold=True, size=14)
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Row 2: Plant section (centered)
        ws.append(["", "", "POTLINE, FUME TREATMENT PLANT"] + [""] * (total_cols - 3))
        ws.merge_cells(f"C2:{merge_end}2")
        ws["C2"].font = Font(bold=True, size=12)
        ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        # Row 3: Report title and date range (centered)
        from_date_obj = datetime.strptime(report_data["meta"]["from_date"], "%Y-%m-%d")
        to_date_obj = datetime.strptime(report_data["meta"]["to_date"], "%Y-%m-%d")
        date_range_str = f"{from_date_obj.strftime('%d-%B-%y')} to {to_date_obj.strftime('%d-%B-%y')}"
        ws.append(["", "", "MONTHLY REPORT"] + [""] * 7 + [f"PERIOD: {date_range_str}"] + [""] * (total_cols - 11))
        ws.merge_cells("C3:J3")
        ws["C3"].font = Font(bold=True, size=12)
        ws["C3"].alignment = Alignment(horizontal="center", vertical="center")
        # PERIOD label starts at column K (11). For short date ranges total_cols
        # can be < 11, which would build an invalid merge (e.g. K3:I3). Clamp the
        # end column so it is never to the left of K.
        merge_date_end = get_column_letter(max(11, min(16, total_cols)))
        ws.merge_cells(f"K3:{merge_date_end}3")
        ws["K3"].font = Font(bold=True, size=11)
        ws["K3"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 20

        # Row 4: Column headers
        # Format day labels as "1st", "2nd", "3rd", etc.
        day_labels = []
        for col_date in report_data["columns"]:
            day_num = int(col_date.split("-")[2])  # Extract day from YYYY-MM-DD
            if day_num in [1, 21, 31]:
                suffix = "st"
            elif day_num in [2, 22]:
                suffix = "nd"
            elif day_num in [3, 23]:
                suffix = "rd"
            else:
                suffix = "th"
            day_labels.append(f"{day_num}{suffix}")
        
        # COLUMN ORDER: Equipment | Sub Equipment | Tag Name | Tag Description | Unit | AVG | MIN | MAX | [days]
        headers_row = ["Equipment", "Sub Equipment", "Tag Name", "Tag Description", "Unit", "AVG", "MIN", "MAX"] + day_labels
        ws.append(headers_row)
        
        # Style header row
        header_fill = PatternFill("solid", fgColor="4472C4")
        summary_fill = PatternFill("solid", fgColor="E7E6E6")
        white_bold = Font(color="FFFFFF", bold=True, size=10)
        for cell in ws[4]:
            if cell.value:
                cell.fill = header_fill
                cell.font = white_bold
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Helper to filter out 'None' string and None/empty values
        def clean_field(value):
            if value is None or value == '' or value == 'None':
                return ''
            return value
        
        # Data rows - Format: Equipment | Sub Equipment | Tag Name | Tag Description | Unit | AVG | MIN | MAX | [days]
        sorted_rows = sorted(
            report_data["rows"],
            key=lambda r: (
                (r.get("group") or "").lower(),
                (r.get("display_label") or r.get("tag_id") or "").lower(),
                (r.get("parameter_unit") or "").lower(),
            ),
        )

        for row in sorted_rows:
            data_row = [
                clean_field(row.get("group")),                      # Equipment
                clean_field(row.get("sub_equipment")),              # Sub Equipment
                row["tag_id"],                                      # Tag Name
                clean_field(row.get("description")) or clean_field(row.get("display_label")) or row["tag_id"],  # Tag Description
                clean_field(row.get("eng_unit")),                   # Unit
                row["avg"],                                         # AVG  ← col 6
                row["min"],                                         # MIN  ← col 7
                row["max"],                                         # MAX  ← col 8
                *row["hourly"],                                     # Daily values ← col 9+
            ]
            ws.append(data_row)

        # Style data rows - NO MERGING, each row independent
        start_data_row = 5
        end_data_row = ws.max_row
        
        for r in range(start_data_row, end_data_row + 1):
            for c in range(1, len(headers_row) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                
                if c >= 6 and isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                # Highlight AVG, MIN, MAX columns (fixed positions 6, 7, 8)
                if c in (6, 7, 8):
                    cell.fill = summary_fill
                # Center align Equipment, Sub Equipment, and Unit columns
                if c in (1, 2, 5):
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # Left align Tag Name and Tag Description
                if c in (3, 4):
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Column widths
        ws.column_dimensions["A"].width = 22   # Equipment
        ws.column_dimensions["B"].width = 18   # Sub Equipment
        ws.column_dimensions["C"].width = 24   # Tag Name
        ws.column_dimensions["D"].width = 40   # Tag Description
        ws.column_dimensions["E"].width = 12   # Unit
        for col in range(6, len(headers_row) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    # ------------------------------------------------------------------ #
    # Shift Report                                                         #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _hour_to_label(h: int) -> str:
        suffix_s = "am" if h < 12 else "pm"
        h12_s = h % 12 or 12
        end = (h + 1) % 24
        suffix_e = "am" if end < 12 else "pm"
        h12_e = end % 12 or 12
        return f"{h12_s} {suffix_s} To {h12_e} {suffix_e}"

    @staticmethod
    def _load_shift(cursor, shift_code: str) -> Dict[str, Any]:
        cursor.execute(
            """
            SELECT id, shift_code, shift_name, start_time, end_time
            FROM historian_meta.shifts
            WHERE shift_code = %s AND is_active = TRUE
            """,
            (shift_code,),
        )
        row = cursor.fetchone()
        if not row:
            raise ValueError(f"Shift '{shift_code}' not found or inactive")
        return row

    @staticmethod
    def _shift_hours(shift_row: Dict[str, Any]) -> List[int]:
        st = shift_row["start_time"]
        et = shift_row["end_time"]
        start_h = st.hour if hasattr(st, "hour") else int(str(st).split(":")[0])
        end_h = et.hour if hasattr(et, "hour") else int(str(et).split(":")[0])
        if end_h > start_h:
            return list(range(start_h, end_h))
        # cross-day (end_h <= start_h)
        return list(range(start_h, 24)) + list(range(0, end_h))

    @staticmethod
    def _is_cross_day(shift_row: Dict[str, Any]) -> bool:
        st = shift_row["start_time"]
        et = shift_row["end_time"]
        start_h = st.hour if hasattr(st, "hour") else int(str(st).split(":")[0])
        end_h = et.hour if hasattr(et, "hour") else int(str(et).split(":")[0])
        return end_h <= start_h

    def build_shift_report(
        self,
        report_date: datetime.date,
        shift_code: str,
        plant: str | list[str],
        area: str | list[str],
        generated_by: str,
        page: int | None = None,
        page_size: int | None = None,
        source_id: str | None = None,
    ) -> Dict[str, Any]:
        if not self.connection:
            raise RuntimeError("Database not connected")

        # Normalize plant and area to lists for multi-select support
        plants = [plant] if isinstance(plant, str) else plant
        areas = [area] if isinstance(area, str) else area

        with self.connection.cursor() as cursor:
            shift_row = self._load_shift(cursor, shift_code)
            ordered_hours = self._shift_hours(shift_row)
            columns = [self._hour_to_label(h) for h in ordered_hours]
            cross_day = self._is_cross_day(shift_row)
            next_date = report_date + timedelta(days=1)
            st = shift_row["start_time"]
            et = shift_row["end_time"]
            start_h = st.hour if hasattr(st, "hour") else int(str(st).split(":")[0])
            end_h = et.hour if hasattr(et, "hour") else int(str(et).split(":")[0])
            

            meta = {
                "company": self.app_config.get("reporting", {}).get("company_name", "BHARAT ALUMINIUM COMPANY LIMITED"),
                "plant": ', '.join(plants),
                "report_title": f"Utility Report Of :- {', '.join(areas)}",
                "date": report_date.isoformat(),
                "shift_code": shift_row["shift_code"],
                "shift_name": shift_row["shift_name"],
                "shift_start": str(shift_row["start_time"]),
                "shift_end": str(shift_row["end_time"]),
                "generated_at": datetime.now().isoformat(),
                "generated_by": generated_by,
            }
            
            template_rows = []

            if not source_id:
                # Build IN clause for multiple areas and plants
                area_placeholders = ",".join(["%s"] * len(areas))
                plant_placeholders = ",".join(["%s"] * len(plants))
                cursor.execute(
                    f"""
                    SELECT
                        vt.s_no,
                        vt.tag_id,
                        vt.display_label,
                        vt.group_name,
                        vt.parameter_unit,
                        vt.plant,
                        vt.area,
                        tm.sub_equipment,
                        tm.description,
                        tm.eng_unit
                    FROM historian_meta.v_report_template_tags vt
                    LEFT JOIN historian_meta.tag_master tm ON vt.tag_id = tm.tag_id
                    WHERE vt.report_type = 'SHIFT'
                      AND vt.plant IN ({plant_placeholders})
                      AND vt.area IN ({area_placeholders})
                      AND vt.template_enabled = TRUE
                      AND vt.tag_enabled = TRUE
                    ORDER BY vt.group_name ASC, vt.s_no ASC
                    """,
                    (*plants, *areas),
                )
                template_rows = cursor.fetchall()

            if not template_rows:
                source_filter = "AND tm.server_progid = %s" if source_id else ""
                area_placeholders = ",".join(["%s"] * len(areas))
                plant_placeholders = ",".join(["%s"] * len(plants))
                fallback_params = (*plants, *areas, source_id) if source_id else (*plants, *areas)
                cursor.execute(
                    f"""
                    SELECT
                        ROW_NUMBER() OVER (ORDER BY tm.equipment, tm.tag_name) AS s_no,
                        tm.tag_id,
                        tm.tag_name AS display_label,
                        tm.equipment AS group_name,
                        COALESCE(tm.eng_unit, tm.data_type, '') AS parameter_unit,
                        tm.plant,
                        tm.area,
                        tm.sub_equipment,
                        tm.description,
                        tm.eng_unit
                    FROM historian_meta.tag_master tm
                    WHERE tm.plant IN ({plant_placeholders})
                      AND tm.area IN ({area_placeholders})
                      {source_filter}
                      AND tm.enabled = TRUE
                      AND COALESCE(tm.include_in_report, TRUE) = TRUE
                      AND tm.tag_id IN (
                          SELECT DISTINCT tag_id FROM historian_raw.v_daily_hourly_agg
                      )
                    ORDER BY tm.equipment, tm.tag_name
                    """,
                    fallback_params,
                )
                template_rows = cursor.fetchall()

            if not template_rows:
                return {
                    "meta": meta,
                    "columns": columns,
                    "rows": [],
                    "pagination": {
                        "page": 1,
                        "page_size": int(page_size) if page_size else 0,
                        "total_rows": 0,
                        "total_pages": 0,
                    },
                }

            tag_ids = [row["tag_id"] for row in template_rows]

            if not cross_day:
                cursor.execute(
                    """
                    SELECT tag_id, local_date, local_hour AS hour, avg_val, max_val, min_val
                    FROM historian_raw.v_daily_hourly_agg
                    WHERE tag_id = ANY(%s)
                      AND local_date = %s
                      AND local_hour >= %s
                      AND local_hour < %s
                    ORDER BY tag_id, local_date, local_hour
                    """,
                    (tag_ids, report_date, start_h, end_h),
                )
            else:
                cursor.execute(
                    """
                    SELECT tag_id, local_date, local_hour AS hour, avg_val, max_val, min_val
                    FROM historian_raw.v_daily_hourly_agg
                    WHERE tag_id = ANY(%s)
                      AND (
                          (local_date = %s AND local_hour >= %s)
                          OR
                          (local_date = %s AND local_hour < %s)
                      )
                    ORDER BY tag_id, local_date, local_hour
                    """,
                    (tag_ids, report_date, start_h, next_date, end_h),
                )
            agg_rows = cursor.fetchall()

        hourly_map: Dict[str, Dict[int, Dict[str, Any]]] = {}
        for row in agg_rows:
            tag_id = row["tag_id"]
            hour = int(row["hour"])
            hourly_map.setdefault(tag_id, {})[hour] = {
                "avg": float(row["avg_val"]) if row["avg_val"] is not None else None,
                "max": float(row["max_val"]) if row["max_val"] is not None else None,
                "min": float(row["min_val"]) if row["min_val"] is not None else None,
            }

        # Helper to filter out 'None' string and None/empty values
        def clean_field(value):
            if value is None or value == '' or value == 'None':
                return ''
            return value
        
        rows_out: List[Dict[str, Any]] = []
        for trow in template_rows:
            tag_id = trow["tag_id"]
            hour_data = hourly_map.get(tag_id, {})
            hourly_values: List[float | None] = []
            all_avg: List[float] = []
            all_max: List[float] = []
            all_min: List[float] = []
            for h in ordered_hours:
                item = hour_data.get(h)
                if item and item["avg"] is not None:
                    hourly_values.append(item["avg"])
                    all_avg.append(item["avg"])
                    if item["max"] is not None:
                        all_max.append(item["max"])
                    if item["min"] is not None:
                        all_min.append(item["min"])
                else:
                    hourly_values.append(None)

            rows_out.append({
                "s_no": trow["s_no"],
                "group": clean_field(trow.get("group_name")),
                "sub_equipment": clean_field(trow.get("sub_equipment")),
                "tag_id": tag_id,
                "description": clean_field(trow.get("description")) or clean_field(trow.get("display_label")) or tag_id,
                "eng_unit": clean_field(trow.get("eng_unit")),
                "parameter_unit": clean_field(trow.get("parameter_unit")),
                "display_label": clean_field(trow.get("display_label")) or tag_id,
                "avg": round(sum(all_avg) / len(all_avg), 2) if all_avg else None,
                "max": round(max(all_max), 2) if all_max else None,
                "min": round(min(all_min), 2) if all_min else None,
                "hourly": hourly_values,
            })

        total_rows = len(rows_out)
        rows_for_response = rows_out
        pagination = {
            "page": 1,
            "page_size": total_rows,
            "total_rows": total_rows,
            "total_pages": 1 if total_rows else 0,
        }

        if page is not None and page_size is not None:
            safe_page_size = max(1, int(page_size))
            total_pages = max(1, (total_rows + safe_page_size - 1) // safe_page_size)
            safe_page = min(max(1, int(page)), total_pages)
            start = (safe_page - 1) * safe_page_size
            rows_for_response = rows_out[start: start + safe_page_size]
            pagination = {
                "page": safe_page,
                "page_size": safe_page_size,
                "total_rows": total_rows,
                "total_pages": total_pages,
            }

        return {
            "meta": meta,
            "columns": columns,
            "rows": rows_for_response,
            "pagination": pagination,
        }

    def export_shift_to_excel(self, report_data: Dict[str, Any], tag_ids: list | None = None) -> BytesIO:
        # Apply tag filter if provided (download-filtered feature)
        if tag_ids:
            report_data = dict(report_data)
            report_data["rows"] = [r for r in report_data["rows"] if r["tag_id"] in tag_ids]

        wb = Workbook()
        ws = wb.active
        ws.title = "Shift Report"

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Add company logo on the left side (Row 1)
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "apex-hmi", "public", "Logo_Company.png")
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.height = 60
            img.width = int(img.width * (60 / img.height))
            ws.add_image(img, "A1")
        
        # Row 1: Company name and plant (centered, leaving space for logo)
        ws.append(["", "", "BHARAT ALUMINIUM COMPANY LIMITED ( PLANT- II )", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws.merge_cells("C1:N1")
        ws["C1"].font = Font(bold=True, size=14)
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Row 2: Plant section (centered)
        ws.append(["", "", "POTLINE, FUME TREATMENT PLANT", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws.merge_cells("C2:N2")
        ws["C2"].font = Font(bold=True, size=12)
        ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        # Row 3: Report title, shift, and date (centered)
        date_obj = datetime.strptime(report_data["meta"]["date"], "%Y-%m-%d")
        date_str = date_obj.strftime("%d-%B-%y")
        shift_name = report_data["meta"]["shift_name"]
        ws.append(["", "", "SHIFT REPORT  (CONTROL ROOM)", "", "", "", "", "", "", "", "", "", f"SHIFT: {shift_name}  |  DATE: {date_str}", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws.merge_cells("C3:K3")
        ws["C3"].font = Font(bold=True, size=12)
        ws["C3"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("M3:P3")
        ws["M3"].font = Font(bold=True, size=11)
        ws["M3"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 20

        # Row 4: Column headers
        # EXACT FORMAT: Equipment | Sub Equipment | Tag Name | Tag Description | Unit | hours | Min | Max | Avg
        hour_labels = []
        for hour_col in report_data["columns"]:
            parts = hour_col.split(" To ")
            if len(parts) == 2:
                hour_labels.append(f"{parts[0].upper()} \n To {parts[1].upper()}")
            else:
                hour_labels.append(hour_col)
        
        # COLUMN ORDER: Equipment | Sub Equipment | Tag Name | Tag Description | Unit | AVG | MIN | MAX | [hours]
        headers_row = ["Equipment", "Sub Equipment", "Tag Name", "Tag Description", "Unit", "AVG", "MIN", "MAX"] + hour_labels
        ws.append(headers_row)
        
        # Style header row
        header_fill = PatternFill("solid", fgColor="4472C4")
        summary_fill = PatternFill("solid", fgColor="E7E6E6")
        white_bold = Font(color="FFFFFF", bold=True, size=10)
        for cell in ws[4]:
            if cell.value:
                cell.fill = header_fill
                cell.font = white_bold
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Helper to filter out 'None' string and None/empty values
        def clean_field(value):
            if value is None or value == '' or value == 'None':
                return ''
            return value
        
        # Data rows - Format: Equipment | Sub Equipment | Tag Name | Tag Description | Unit | AVG | MIN | MAX | [hours]
        sorted_rows = sorted(
            report_data["rows"],
            key=lambda r: (
                (r.get("group") or "").lower(),
                (r.get("display_label") or r.get("tag_id") or "").lower(),
                (r.get("parameter_unit") or "").lower(),
            ),
        )

        for row in sorted_rows:
            data_row = [
                clean_field(row.get("group")),                      # Equipment
                clean_field(row.get("sub_equipment")),              # Sub Equipment
                row["tag_id"],                                      # Tag Name
                clean_field(row.get("description")) or clean_field(row.get("display_label")) or row["tag_id"],  # Tag Description
                clean_field(row.get("eng_unit")),                   # Unit
                row["avg"],                                         # AVG  ← col 6
                row["min"],                                         # MIN  ← col 7
                row["max"],                                         # MAX  ← col 8
                *row["hourly"],                                     # Hourly values ← col 9+
            ]
            ws.append(data_row)

        # Style data rows - NO MERGING, each row independent
        start_data_row = 5
        end_data_row = ws.max_row
        
        for r in range(start_data_row, end_data_row + 1):
            for c in range(1, len(headers_row) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                
                if c >= 6 and isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                # Highlight AVG, MIN, MAX columns (fixed positions 6, 7, 8)
                if c in (6, 7, 8):
                    cell.fill = summary_fill
                # Center align Equipment, Sub Equipment, and Unit columns
                if c in (1, 2, 5):
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # Left align Tag Name and Tag Description
                if c in (3, 4):
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Column widths - updated for new structure
        ws.column_dimensions["A"].width = 22  # Equipment
        ws.column_dimensions["B"].width = 18  # Sub Equipment
        ws.column_dimensions["C"].width = 24  # Tag Name
        ws.column_dimensions["D"].width = 40  # Tag Description
        ws.column_dimensions["E"].width = 12  # Unit
        for col in range(6, len(headers_row) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out
