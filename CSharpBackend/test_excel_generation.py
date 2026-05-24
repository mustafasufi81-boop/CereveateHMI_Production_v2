import requests
import sys
from datetime import datetime

print("=" * 80)
print("EXCEL GENERATION TEST - Step by Step Debugging")
print("=" * 80)

# Step 1: Login
print("\n[STEP 1] Logging in...")
login_response = requests.post(
    "http://localhost:6001/api/auth/login",
    json={"username": "Mustafa", "password": "Admin@123"}
)

if login_response.status_code != 200:
    print(f"❌ Login failed: {login_response.status_code}")
    print(login_response.text)
    sys.exit(1)

token = login_response.json()["token"]
print(f"✅ Login successful, token: {token[:30]}...")

# Step 2: Get JSON report data
print("\n[STEP 2] Fetching JSON report data...")
headers = {"Authorization": f"Bearer {token}"}
json_url = "http://localhost:6001/api/reports/daily?date=2026-05-18&plant=FTP-1&area=POTLINE"

json_response = requests.get(json_url, headers=headers)
if json_response.status_code != 200:
    print(f"❌ JSON request failed: {json_response.status_code}")
    print(json_response.text)
    sys.exit(1)

json_data = json_response.json()
rows = json_data.get("rows", [])
print(f"✅ JSON data received: {len(rows)} rows")

if len(rows) > 0:
    print("\n[STEP 2a] Inspecting first row data:")
    first_row = rows[0]
    print(f"  tag_id: '{first_row.get('tag_id')}'")
    print(f"  group: '{first_row.get('group')}'")
    print(f"  sub_equipment: '{first_row.get('sub_equipment')}'")
    print(f"  description: '{first_row.get('description')}'")
    print(f"  eng_unit: '{first_row.get('eng_unit')}'")
    print(f"  display_label: '{first_row.get('display_label')}'")
    print(f"  parameter_unit: '{first_row.get('parameter_unit')}'")
    print(f"  hourly values count: {len(first_row.get('hourly', []))}")
    print(f"  min: {first_row.get('min')}, max: {first_row.get('max')}, avg: {first_row.get('avg')}")
    
    # Check if any critical fields are empty/None
    critical_empty = []
    if not first_row.get('sub_equipment'):
        critical_empty.append('sub_equipment')
    if not first_row.get('description'):
        critical_empty.append('description')
    if not first_row.get('eng_unit'):
        critical_empty.append('eng_unit')
    
    if critical_empty:
        print(f"\n  ⚠️ WARNING: These fields are empty: {', '.join(critical_empty)}")
    else:
        print(f"\n  ✅ All critical fields populated!")
else:
    print("❌ No rows in JSON data!")
    sys.exit(1)

# Step 3: Download Excel file
print("\n[STEP 3] Downloading Excel file...")
excel_url = "http://localhost:6001/api/reports/daily/export?date=2026-05-18&plant=FTP-1&area=POTLINE"

excel_response = requests.get(excel_url, headers=headers)
if excel_response.status_code != 200:
    print(f"❌ Excel download failed: {excel_response.status_code}")
    print(excel_response.text)
    sys.exit(1)

# Save to file
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = f"test_excel_report_{timestamp}.xlsx"

with open(output_file, "wb") as f:
    f.write(excel_response.content)

file_size = len(excel_response.content)
print(f"✅ Excel file saved: {output_file} ({file_size} bytes)")

# Step 4: Verify Excel contents using openpyxl
print("\n[STEP 4] Reading Excel file to verify contents...")
try:
    from openpyxl import load_workbook
    
    wb = load_workbook(output_file)
    ws = wb.active
    
    print(f"  Worksheet: {ws.title}")
    print(f"  Total rows: {ws.max_row}")
    print(f"  Total columns: {ws.max_column}")
    
    # Check header row (row 3)
    print("\n[STEP 4a] Header row (row 3):")
    header_row = []
    for cell in ws[3]:
        if cell.value:
            header_row.append(str(cell.value).replace('\n', ' '))
    print(f"  Headers: {' | '.join(header_row[:10])}...")
    
    # Check if we have expected headers
    expected_headers = ["Equipment", "Sub Equipment", "Tag Name", "Tag Description", "Unit"]
    for i, expected in enumerate(expected_headers, start=1):
        actual = ws.cell(3, i).value
        if actual and expected.lower() in str(actual).lower():
            print(f"  ✅ Column {i}: {actual}")
        else:
            print(f"  ❌ Column {i}: Expected '{expected}', got '{actual}'")
    
    # Check first data row (row 4)
    if ws.max_row >= 4:
        print("\n[STEP 4b] First data row (row 4):")
        row4_data = []
        for col in range(1, 6):  # First 5 columns
            cell_value = ws.cell(4, col).value
            row4_data.append(str(cell_value) if cell_value is not None else "")
        
        print(f"  Equipment (A4): '{row4_data[0]}'")
        print(f"  Sub Equipment (B4): '{row4_data[1]}'")
        print(f"  Tag Name (C4): '{row4_data[2]}'")
        print(f"  Tag Description (D4): '{row4_data[3]}'")
        print(f"  Unit (E4): '{row4_data[4]}'")
        
        # Check if critical fields are blank
        blank_fields = []
        if not row4_data[1]:  # Sub Equipment
            blank_fields.append("Sub Equipment (B4)")
        if not row4_data[3]:  # Tag Description
            blank_fields.append("Tag Description (D4)")
        if not row4_data[4]:  # Unit
            blank_fields.append("Unit (E4)")
        
        if blank_fields:
            print(f"\n  ❌ BLANK FIELDS FOUND: {', '.join(blank_fields)}")
            print("\n  🔍 DIAGNOSIS:")
            print("     The Excel file has blank cells even though JSON has data.")
            print("     This means the Excel generation code is NOT using the row data correctly.")
        else:
            print(f"\n  ✅ ALL FIELDS POPULATED IN EXCEL!")
    else:
        print("\n  ❌ No data rows in Excel file!")
    
    wb.close()
    
except ImportError:
    print("  ⚠️ openpyxl not installed, skipping Excel verification")
    print("  Install with: pip install openpyxl")
except Exception as e:
    print(f"  ❌ Error reading Excel: {e}")

print("\n" + "=" * 80)
print("TEST COMPLETE")
print(f"Excel file saved as: {output_file}")
print("Open this file manually to verify the data.")
print("=" * 80)
