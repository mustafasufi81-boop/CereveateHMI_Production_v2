"""
Analyze Tag_master_Details_System_Upload.xlsx vs existing historian_meta.tag_master
Produces:
  - tags_already_in_db.json   (already exist, NO changes needed)
  - tags_to_insert.json       (new tags from Excel not yet in DB)
  - insert_new_tags.sql       (ready-to-run SQL, uses ON CONFLICT DO NOTHING)

DOES NOT TOUCH THE DATABASE - READ ONLY ANALYSIS
"""

import psycopg2
from psycopg2.extras import RealDictCursor
import json
import openpyxl

# ── DB connection ──────────────────────────────────────────────────────────────
conn = psycopg2.connect(
    host='localhost', port=5432,
    dbname='Automation_DB', user='cereveate', password='cereveate@222'
)
cur = conn.cursor(cursor_factory=RealDictCursor)

# ── Load existing tag_ids from DB ──────────────────────────────────────────────
cur.execute("SELECT tag_id FROM historian_meta.tag_master")
existing_ids = {r['tag_id'] for r in cur.fetchall()}
conn.close()
print(f"DB currently has {len(existing_ids)} tags")

# ── Data type mapping: Excel → DB ──────────────────────────────────────────────
DTYPE_MAP = {
    'REAL':   'double',
    'DINT':   'integer',
    'INT':    'integer',
    'BOOL':   'boolean',
    'SINT':   'integer',
    'LINT':   'integer',
    'STRING': 'string',
}

# ── Load Excel ─────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook('Tag_master_Details_System_Upload.xlsx')
ws = wb['Tag Master Data']

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print(f"Excel columns: {headers}")

already_in_db = []
to_insert     = []

for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(v is not None for v in row):
        continue

    r = dict(zip(headers, row))

    tag_id   = r.get('Tag Id (M)')
    tag_name = r.get('Tag Name (M)')
    if not tag_id:
        continue

    # Map data type
    excel_dtype = str(r.get('Data Type (M)') or 'REAL').strip().upper()
    db_dtype    = DTYPE_MAP.get(excel_dtype, 'double')

    # Alarm enabled flag
    alarm_enabled = str(r.get('Alarm Enabled') or '').strip().lower() == 'yes'

    # Alarm priority – default 3 if None or non-numeric
    try:
        alarm_priority = int(r.get('Alarm Priority') or 3)
    except (ValueError, TypeError):
        alarm_priority = 3

    # Equipment criticality
    try:
        eq_crit = int(r.get('Equipment Criticality') or 3)
    except (ValueError, TypeError):
        eq_crit = 3

    # Alarm limits
    def safe_float(v):
        try:
            return float(v) if v is not None else None
        except (ValueError, TypeError):
            return None

    alarm_h  = safe_float(r.get('Alarm High Limit'))
    alarm_hh = safe_float(r.get('Alarm HH Limit'))
    alarm_l  = safe_float(r.get('Alarm Low Limit'))
    alarm_ll = safe_float(r.get('Alarm LL Limit'))

    # Logging interval
    try:
        log_interval = int(r.get('Logging Interval ms') or 1000)
    except (ValueError, TypeError):
        log_interval = 1000

    # PLC details from Excel + known ControlLogix defaults
    plc_ip = r.get('PLC IP Address') or '192.168.1.11'

    record = {
        "tag_id":                  tag_id,
        "tag_name":                tag_name,
        "description":             r.get('Tag Description (M)'),
        "plant":                   r.get('Plant (M)'),
        "area":                    r.get('Area (M)'),
        "equipment":               r.get('Equipment (M)'),
        "sub_equipment":           r.get('Sub Equipment'),
        "components":              r.get('Component (M)'),
        "data_type":               db_dtype,
        "eng_unit":                r.get('Unit (M)'),
        "alarm_enabled":           alarm_enabled,
        "alarm_h_limit":           alarm_h,
        "alarm_hh_limit":          alarm_hh,
        "alarm_l_limit":           alarm_l,
        "alarm_ll_limit":          alarm_ll,
        "alarm_priority":          alarm_priority,
        "alarm_deadband":          1.0,
        "equipment_criticality":   eq_crit,
        "db_logging_interval_ms":  log_interval,
        "logging_interval_ms":     log_interval,
        "deadband_enabled":        False,
        "deadband_value":          None,
        "enabled":                 True,
        "db_table_name":           "historian_raw.historian_timeseries",
        "server_progid":           "Rockwel_PLC_001",
        "plc_ip_address":          plc_ip,
        "plc_port":                44818,
        "plc_protocol":            "Rockwell",
        "plc_slot":                None,
        "plc_path":                "1,0",
        "plc_type":                r.get('PLC Type') or 'ControlLogix',
        "plc_timeout_ms":          3000,
        "plc_polling_interval_ms": 1000,
        "include_in_report":       True,
        "report_flag":             True,
        "created_by":              "excel_import",
    }

    if tag_id in existing_ids:
        already_in_db.append(record)
    else:
        to_insert.append(record)

# ── Save JSONs ─────────────────────────────────────────────────────────────────
with open('tags_already_in_db.json', 'w') as f:
    json.dump(already_in_db, f, indent=2)

with open('tags_to_insert.json', 'w') as f:
    json.dump(to_insert, f, indent=2)

# ── Generate SQL ───────────────────────────────────────────────────────────────
def sql_val(v):
    if v is None:
        return 'NULL'
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, (int, float)):
        return str(v)
    # escape single quotes
    return "'" + str(v).replace("'", "''") + "'"

sql_lines = [
    "-- AUTO-GENERATED: insert new tags from Tag_master_Details_System_Upload.xlsx",
    "-- Uses ON CONFLICT DO NOTHING - safe to run multiple times",
    "-- REVIEW this file before running!\n",
    "INSERT INTO historian_meta.tag_master (",
    "    tag_id, tag_name, description, plant, area, equipment,",
    "    sub_equipment, components, data_type, eng_unit,",
    "    alarm_enabled, alarm_h_limit, alarm_hh_limit, alarm_l_limit, alarm_ll_limit,",
    "    alarm_priority, alarm_deadband, equipment_criticality,",
    "    db_logging_interval_ms, logging_interval_ms,",
    "    deadband_enabled, deadband_value, enabled, db_table_name,",
    "    server_progid, plc_ip_address, plc_port, plc_protocol,",
    "    plc_slot, plc_path, plc_type, plc_timeout_ms, plc_polling_interval_ms,",
    "    include_in_report, report_flag, created_by",
    ") VALUES"
]

rows_sql = []
for t in to_insert:
    rows_sql.append(
        f"    ({sql_val(t['tag_id'])}, {sql_val(t['tag_name'])}, {sql_val(t['description'])}, "
        f"{sql_val(t['plant'])}, {sql_val(t['area'])}, {sql_val(t['equipment'])},\n"
        f"     {sql_val(t['sub_equipment'])}, {sql_val(t['components'])}, {sql_val(t['data_type'])}, {sql_val(t['eng_unit'])},\n"
        f"     {sql_val(t['alarm_enabled'])}, {sql_val(t['alarm_h_limit'])}, {sql_val(t['alarm_hh_limit'])}, "
        f"{sql_val(t['alarm_l_limit'])}, {sql_val(t['alarm_ll_limit'])},\n"
        f"     {sql_val(t['alarm_priority'])}, {sql_val(t['alarm_deadband'])}, {sql_val(t['equipment_criticality'])},\n"
        f"     {sql_val(t['db_logging_interval_ms'])}, {sql_val(t['logging_interval_ms'])},\n"
        f"     {sql_val(t['deadband_enabled'])}, {sql_val(t['deadband_value'])}, {sql_val(t['enabled'])}, {sql_val(t['db_table_name'])},\n"
        f"     {sql_val(t['server_progid'])}, {sql_val(t['plc_ip_address'])}, {sql_val(t['plc_port'])}, {sql_val(t['plc_protocol'])},\n"
        f"     {sql_val(t['plc_slot'])}, {sql_val(t['plc_path'])}, {sql_val(t['plc_type'])}, "
        f"{sql_val(t['plc_timeout_ms'])}, {sql_val(t['plc_polling_interval_ms'])},\n"
        f"     {sql_val(t['include_in_report'])}, {sql_val(t['report_flag'])}, {sql_val(t['created_by'])})"
    )

sql_lines.append(",\n".join(rows_sql))
sql_lines.append("ON CONFLICT (tag_id) DO NOTHING;\n")
sql_lines.append(f"\n-- Verify: SELECT COUNT(*) FROM historian_meta.tag_master WHERE plant = 'FTP-1';")

with open('insert_new_tags.sql', 'w', encoding='utf-8') as f:
    f.write("\n".join(sql_lines))

# ── Summary ────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print(f"  EXCEL total tags   : {len(already_in_db) + len(to_insert)}")
print(f"  Already in DB      : {len(already_in_db)}  →  tags_already_in_db.json")
print(f"  NEW (to insert)    : {len(to_insert)}  →  tags_to_insert.json")
print(f"  SQL file generated : insert_new_tags.sql")
print("=" * 60)

print("\n  ALREADY IN DB (no action needed):")
for t in already_in_db:
    print(f"    ✅  {t['tag_id']:20s}  {t['description']}")

print("\n  NEW TAGS TO INSERT:")
for t in to_insert:
    print(f"    ➕  {t['tag_id']:20s}  {t['description']}")

print("\nDONE — nothing written to DB. Review files then run insert_new_tags.sql")
