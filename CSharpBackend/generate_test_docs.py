"""
Generates two Excel test function documents:
  1. TEST_FUNCTIONS_HMI.xlsx       — HMI side: Alarms, Trends, Reports, Auth, Admin
  2. TEST_FUNCTIONS_SERVER.xlsx    — Server side: OPC, Historian DB, TimescaleDB, PLC drivers
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── COLOUR PALETTE ──────────────────────────────────────────────
C_HEADER_DARK   = "1F3864"   # dark navy  — sheet header row
C_HEADER_MID    = "2E75B6"   # blue       — section header
C_HEADER_LIGHT  = "D6E4F0"   # light blue — column header row
C_PASS          = "C6EFCE"   # green      — expected result cell
C_FAIL_HINT     = "FCE4D6"   # orange     — known failure/warning row
C_ALT           = "F2F7FB"   # very light — alternate data row
C_WHITE         = "FFFFFF"
C_GOLD          = "FFC000"   # gold       — section label
C_DARK_TEXT     = "1F3864"
C_LIGHT_TEXT    = "FFFFFF"


def thin_border():
    s = Side(style="thin", color="B0B8C1")
    return Border(left=s, right=s, top=s, bottom=s)


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def hdr_font(size=11, bold=True, color=C_LIGHT_TEXT):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def body_font(size=10, bold=False, color=C_DARK_TEXT):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def wrap_center():
    return Alignment(wrap_text=True, vertical="center", horizontal="center")


def wrap_left():
    return Alignment(wrap_text=True, vertical="center", horizontal="left")


# ─── HELPERS ─────────────────────────────────────────────────────

def write_cover(ws, title, subtitle, scope_lines):
    """Write a cover / summary tab."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30

    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 30

    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = title
    c.font = Font(name="Calibri", size=22, bold=True, color=C_LIGHT_TEXT)
    c.fill = make_fill(C_HEADER_DARK)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")

    ws.merge_cells("B3:C3")
    c = ws["B3"]
    c.value = subtitle
    c.font = Font(name="Calibri", size=12, bold=False, color=C_LIGHT_TEXT)
    c.fill = make_fill(C_HEADER_MID)
    c.alignment = wrap_left()

    row = 5
    for line in scope_lines:
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"B{row}:C{row}")
        c = ws[f"B{row}"]
        c.value = line
        c.font = body_font(10)
        c.alignment = wrap_left()
        row += 1


def write_sheet_header(ws, title, col_headers, col_widths):
    """Write title row + column header row. Returns next row number."""
    ws.sheet_view.showGridLines = False

    span = len(col_headers)
    last_col = get_column_letter(span)

    # Title row
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = hdr_font(13)
    c.fill = make_fill(C_HEADER_DARK)
    c.alignment = wrap_center()
    ws.row_dimensions[1].height = 30

    # Column header row
    for i, (hdr, w) in enumerate(zip(col_headers, col_widths), start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = w
        cell = ws.cell(row=2, column=i, value=hdr)
        cell.font = hdr_font(10)
        cell.fill = make_fill(C_HEADER_LIGHT)
        cell.alignment = wrap_center()
        cell.border = thin_border()
        cell.font = Font(name="Calibri", size=10, bold=True, color=C_DARK_TEXT)
    ws.row_dimensions[2].height = 28

    return 3   # next data row


def write_section_label(ws, row, label, span):
    """Write a coloured section divider row."""
    last_col = get_column_letter(span)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws[f"A{row}"]
    c.value = f"  {label}"
    c.font = Font(name="Calibri", size=10, bold=True, color=C_DARK_TEXT)
    c.fill = make_fill(C_GOLD)
    c.alignment = wrap_left()
    ws.row_dimensions[row].height = 20
    return row + 1


def write_data_row(ws, row, values, alt=False, highlight=None):
    """Write a data row. highlight='pass'|'warn'|None."""
    bg = C_ALT if alt else C_WHITE
    if highlight == "pass":
        bg = C_PASS
    elif highlight == "warn":
        bg = C_FAIL_HINT

    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = make_fill(bg)
        cell.font = body_font(9)
        cell.border = thin_border()
        cell.alignment = wrap_left()
    ws.row_dimensions[row].height = 42
    return row + 1


# ═══════════════════════════════════════════════════════════════════
#  HMI WORKBOOK
# ═══════════════════════════════════════════════════════════════════

def build_hmi_workbook():
    wb = openpyxl.Workbook()

    # ── COVER ────────────────────────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "COVER"
    write_cover(ws_cover,
        "HMI Application — Test Function Document",
        "Cereveate OPC DA Historian Platform  |  May 2026",
        [
            "Scope: React/Vite HMI (port 8090) + Flask API (port 6001)",
            "",
            "Modules Covered:",
            "  1. Authentication & MFA",
            "  2. Dashboard / Live Tag Values",
            "  3. Alarm Management",
            "  4. Trend / Historical Data",
            "  5. Reports (Daily / Shift / Monthly)",
            "  6. Admin Panel (Users, Roles, Permissions)",
            "  7. Asset Browser / Hierarchy",
            "  8. Audit Trail",
            "",
            "Test Environment:",
            "  HMI URL     :  http://localhost:8090",
            "  API Base    :  http://localhost:6001",
            "  Login       :  Mustafa  /  Admin@123",
            "  DB          :  Automation_DB  @  localhost:5432",
            "",
            "Legend:",
            "  GREEN row  = Expected PASS result",
            "  ORANGE row = Expected edge case / known behaviour",
            "  Status column: PASS / FAIL / SKIP",
        ]
    )

    # ── 1. AUTH ──────────────────────────────────────────────────
    ws = wb.create_sheet("1_Authentication")
    COLS = ["TC#", "Module", "Test Case Name", "Pre-Condition",
            "Test Steps", "Test Input / Values", "Expected Result", "Actual Result", "Status", "Notes"]
    WIDTHS = [6, 16, 28, 25, 40, 30, 35, 25, 10, 20]
    row = write_sheet_header(ws, "Module 1 — Authentication & MFA", COLS, WIDTHS)

    row = write_section_label(ws, row, "LOGIN", len(COLS))
    tests = [
        ("TC-A01","Auth","Valid login","App running, user exists","1. Go to http://localhost:8090\n2. Enter username & password\n3. Click Login","Username: Mustafa\nPassword: Admin@123","Redirect to MFA screen OR dashboard if MFA disabled","","",""),
        ("TC-A02","Auth","Invalid password","App running","1. Enter wrong password\n2. Click Login","Username: Mustafa\nPassword: WrongPass","Error message: 'Invalid credentials'. No redirect.","","",""),
        ("TC-A03","Auth","Empty fields","App running","1. Leave username blank\n2. Click Login","Username: (blank)\nPassword: (blank)","Form validation error — fields highlighted","","",""),
        ("TC-A04","Auth","Non-existent user","App running","1. Enter unknown username","Username: unknown123\nPassword: anything","Error: 'Invalid credentials'","","",""),
        ("TC-A05","MFA","MFA code — valid","Login successful, MFA enabled","1. After login enter correct 6-digit TOTP","TOTP from authenticator app","Redirect to dashboard","","",""),
        ("TC-A06","MFA","MFA code — wrong","Login successful","Enter wrong 6-digit code","123456 (wrong)","Error: 'Invalid MFA code'. Stay on MFA screen.","","",""),
        ("TC-A07","Auth","Session expiry","Logged in > 8h","Wait for session timeout OR check JWT expiry","N/A","Redirect to login page with session expired message","","",""),
        ("TC-A08","Auth","Logout","Logged in","Click logout button","N/A","Redirect to login, JWT cleared, API returns 401","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "API AUTH ENDPOINTS", len(COLS))
    api_tests = [
        ("TC-A09","API","POST /api/auth/login — valid","Flask running","POST with valid JSON body","Body: {\"username\":\"Mustafa\",\"password\":\"Admin@123\"}","HTTP 200, returns {token, user_id, mfa_required}","","",""),
        ("TC-A10","API","POST /api/auth/login — invalid","Flask running","POST with wrong creds","Body: {\"username\":\"x\",\"password\":\"x\"}","HTTP 401, {error: 'Invalid credentials'}","","",""),
        ("TC-A11","API","GET /api/auth/me — no token","Flask running","GET without Authorization header","Headers: (none)","HTTP 401 Unauthorized","","",""),
        ("TC-A12","API","GET /api/auth/me — valid token","Logged in, have JWT","GET with Bearer token","Headers: Authorization: Bearer <token>","HTTP 200, returns current user details","","",""),
    ]
    for i, t in enumerate(api_tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 2. DASHBOARD / LIVE ───────────────────────────────────────
    ws = wb.create_sheet("2_Dashboard_LiveTags")
    row = write_sheet_header(ws, "Module 2 — Dashboard & Live Tag Values", COLS, WIDTHS)

    row = write_section_label(ws, row, "LIVE TAG DISPLAY", len(COLS))
    tests = [
        ("TC-D01","Dashboard","Tags load on dashboard","OPC connected, tags configured","1. Login\n2. Navigate to Dashboard","N/A","Tag values displayed — updating every ~1 second","","",""),
        ("TC-D02","Dashboard","Tag value updates","OPC connected","1. Observe a tag value (e.g. Random.Real4)\n2. Wait 3 seconds","N/A","Value changes on screen — not frozen","","",""),
        ("TC-D03","Dashboard","Bad quality tag display","OPC disconnected or tag offline","1. Disconnect OPC\n2. Observe tags","N/A","Tags show 'Bad' quality indicator, not fake values","","",""),
        ("TC-D04","API","GET /api/opc/values","OPC connected","Call endpoint from browser/curl","GET http://localhost:5001/api/opc/values","HTTP 200, JSON array of {tagId, value, quality, timestamp}","","",""),
        ("TC-D05","API","GET /api/opc/values — OPC down","OPC disconnected","Call endpoint","GET http://localhost:5001/api/opc/values","HTTP 200 but quality='Bad' on all tags — NOT fake values","","","CRITICAL: no simulation"),
        ("TC-D06","SignalR","Real-time push","OPC connected, browser open","1. Open dashboard\n2. Check browser DevTools Network WS tab","ws://localhost:5001/opcHub","WebSocket connection active, TagValuesUpdated messages arriving","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "API — OPC TAG ENDPOINTS", len(COLS))
    tests2 = [
        ("TC-D07","API","GET /api/opc/connections","OPC backend running","GET request","GET http://localhost:5001/api/opc/connections","HTTP 200, list of connections with status","","",""),
        ("TC-D08","API","GET /api/opc/tags","OPC connected","GET request","GET http://localhost:5001/api/opc/tags","HTTP 200, array of configured tag IDs","","",""),
    ]
    for i, t in enumerate(tests2):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 3. ALARMS ─────────────────────────────────────────────────
    ws = wb.create_sheet("3_Alarms")
    row = write_sheet_header(ws, "Module 3 — Alarm Management", COLS, WIDTHS)

    row = write_section_label(ws, row, "ACTIVE ALARMS", len(COLS))
    tests = [
        ("TC-AL01","Alarms","Active alarms load","Flask running, alarms in DB","1. Navigate to Alarms page\n2. Click Active tab","N/A","Table shows active alarms with: Tag ID, Description, Severity, Timestamp, ACK status","","",""),
        ("TC-AL02","Alarms","Alarm count badge","Active alarms exist","1. Check navigation sidebar","N/A","Alarm icon shows count badge matching active alarm count","","",""),
        ("TC-AL03","API","GET /api/alarms/active","Flask running","GET with auth token","GET http://localhost:6001/api/alarms/active\nHeaders: Authorization: Bearer <token>","HTTP 200, {success:true, alarms:[...], count:N}","","",""),
        ("TC-AL04","API","GET /api/alarms/stats","Flask running","GET request","GET http://localhost:6001/api/alarms/stats","HTTP 200, counts by severity: {critical, high, medium, low}","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "ALARM ACKNOWLEDGE / CLEAR", len(COLS))
    tests2 = [
        ("TC-AL05","Alarms","Acknowledge single alarm","Active unacked alarm exists","1. Click ACK button on alarm row\n2. Enter comment","Alarm ID: from active list\nComment: 'Acknowledged by operator'","Alarm moves to ACK state. Audit trail entry created.","","",""),
        ("TC-AL06","API","POST /api/alarms/acknowledge/<id>","Active alarm in DB","POST with body","POST http://localhost:6001/api/alarms/acknowledge/1\nBody: {comment:'test ack'}","HTTP 200, {success:true}. DB updated: ack_status=true","","",""),
        ("TC-AL07","Alarms","Clear alarm","ACKed alarm exists","1. Click Clear on alarm","Alarm ID: ACKed alarm","Alarm removed from active list. History updated.","","",""),
        ("TC-AL08","API","POST /api/alarms/clear/<id>","ACKed alarm in DB","POST request","POST http://localhost:6001/api/alarms/clear/1","HTTP 200. Alarm cleared from active table.","","",""),
        ("TC-AL09","Alarms","Acknowledge by tag ID","Tag with active alarm","POST acknowledge by tag","Body: {tag_id:'Random.Real4', comment:'test'}","All alarms for that tag acknowledged","","",""),
    ]
    for i, t in enumerate(tests2):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "ALARM HISTORY / AUDIT", len(COLS))
    tests3 = [
        ("TC-AL10","Alarms","Alarm history loads","Historical alarms in DB","1. Navigate to Alarm History tab","N/A","Table shows past alarms with timestamps, operator actions","","",""),
        ("TC-AL11","API","GET /api/alarms/history","Flask running","GET with date filter","GET /api/alarms/history?start=2026-05-01&end=2026-05-20","HTTP 200, array of historical alarm events","","",""),
        ("TC-AL12","Alarms","Alarm audit trail","Alarm with audit entries","1. Click audit icon on alarm","Alarm ID with history","Popup/panel shows: created, ack, cleared with operator & time","","",""),
        ("TC-AL13","API","GET /api/alarms/audit/<id>","Alarm with audit","GET request","GET /api/alarms/audit/1","HTTP 200, array of audit events for that alarm","","",""),
    ]
    for i, t in enumerate(tests3):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "ALARM SUPPRESSION", len(COLS))
    tests4 = [
        ("TC-AL14","Alarms","Suppress alarm","Active alarm exists","1. Click Suppress\n2. Set end time","End time: +1 hour from now","Alarm suppressed. Shown in suppressed list. No new notification during window.","","",""),
        ("TC-AL15","API","POST /api/alarms/suppress/<id>","Active alarm","POST with duration","Body: {duration_minutes:60, reason:'Maintenance window'}","HTTP 200. Alarm status='suppressed'","","",""),
        ("TC-AL16","API","GET /api/alarms/suppressed","Flask running","GET request","GET /api/alarms/suppressed","HTTP 200, list of currently suppressed alarms","","",""),
    ]
    for i, t in enumerate(tests4):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 4. TRENDS ─────────────────────────────────────────────────
    ws = wb.create_sheet("4_Trends_Historical")
    row = write_sheet_header(ws, "Module 4 — Trends & Historical Data", COLS, WIDTHS)

    row = write_section_label(ws, row, "HISTORIAN TREND VIEWER (Backup Page)", len(COLS))
    tests = [
        ("TC-TR01","Trends","Historian page loads","Historian DB populated","1. Navigate to Historian tab","N/A","Page loads, tag selector and date picker visible","","",""),
        ("TC-TR02","Trends","Select tag + date range","Historian has data","1. Select tag: Random.Real4\n2. Set range: last 24h\n3. Click Load","Tag: Random.Real4\nStart: now-24h\nEnd: now","Chart renders with data points, X-axis = time, Y-axis = value","","",""),
        ("TC-TR03","Trends","No data range","Empty period","1. Select date range with no data","Start: 2020-01-01\nEnd: 2020-01-02","Chart shows 'No data available' message — NOT error crash","","",""),
        ("TC-TR04","Trends","Multiple tags on same chart","2+ tags in DB","1. Select 2 tags\n2. Load chart","Tags: Random.Real4, Random.Int2","Two lines on chart, different colours, legend shown","","",""),
        ("TC-TR05","API","GET /api/historical/trend","DB running","GET with params","GET http://localhost:6001/api/historical/trend?tag_id=Random.Real4&start=2026-05-19T00:00:00&end=2026-05-20T00:00:00","HTTP 200, {data:[{time, value},...]}","","",""),
        ("TC-TR06","API","TimescaleDB hourly agg","ts_hourly_agg populated","Query agg view","SELECT * FROM historian_raw.ts_hourly_agg WHERE tag_id='Random.Real4' LIMIT 5","Returns rows with bucket, avg_value, min_value, max_value, sample_count","","","DB level test"),
        ("TC-TR07","Trends","Export trend data","Data loaded in chart","1. Click Export CSV button","N/A","CSV downloaded with columns: timestamp, tag_id, value","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "HISTORIAN DB DIRECT QUERIES", len(COLS))
    tests2 = [
        ("TC-TR08","DB","Row count check","DB populated","Run SQL","SELECT count(*) FROM historian_raw.historian_timeseries","Returns ≥ 15,000,000","","",""),
        ("TC-TR09","DB","Chunk count","TimescaleDB active","Run SQL","SELECT count(*) FROM timescaledb_information.chunks WHERE hypertable_name='historian_timeseries'","Returns ≥ 5 chunks","","",""),
        ("TC-TR10","DB","Compressed chunks","Compression ran","Run SQL","SELECT count(*) FROM timescaledb_information.chunks WHERE hypertable_name='historian_timeseries' AND is_compressed=true","Returns ≥ 1 compressed chunk","","",""),
        ("TC-TR11","DB","Hourly agg data","ts_hourly_agg exists","Run SQL","SELECT count(*) FROM historian_raw.ts_hourly_agg","Returns ≥ 6000 rows","","",""),
    ]
    for i, t in enumerate(tests2):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 5. REPORTS ────────────────────────────────────────────────
    ws = wb.create_sheet("5_Reports")
    row = write_sheet_header(ws, "Module 5 — Reports (Daily / Shift / Monthly)", COLS, WIDTHS)

    row = write_section_label(ws, row, "DAILY REPORT", len(COLS))
    tests = [
        ("TC-R01","Reports","Daily report page loads","Flask running, templates seeded","1. Navigate to Reports > Daily","N/A","Date picker + plant/area dropdowns visible","","",""),
        ("TC-R02","Reports","Generate daily report","Report templates in DB","1. Select date: 2026-05-19\n2. Select plant + area\n3. Click Generate","Date: 2026-05-19\nPlant: Plant001\nArea: AreaA","Report table renders with tag rows, hourly averages, min/max","","",""),
        ("TC-R03","API","GET /api/reports/daily","Flask running","GET with params","GET /api/reports/daily?date=2026-05-19&plant=Plant001&area=AreaA\nHeaders: Bearer token","HTTP 200, JSON report data","","",""),
        ("TC-R04","Reports","Export daily report Excel","Daily report loaded","1. Click Export Excel button","N/A","Excel file downloaded — .xlsx with report data","","",""),
        ("TC-R05","API","GET /api/reports/daily/export","Flask running","GET export endpoint","GET /api/reports/daily/export?date=2026-05-19&plant=Plant001&area=AreaA","HTTP 200, Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "SHIFT REPORT", len(COLS))
    tests2 = [
        ("TC-R06","Reports","Shift report page loads","Flask running","Navigate to Reports > Shift","N/A","Shift selector (Morning/Evening/Night), date picker visible","","",""),
        ("TC-R07","Reports","Generate shift report","Shift data exists in DB","1. Select date + shift\n2. Click Generate","Date: 2026-05-19\nShift: Morning (06:00-14:00)","Report renders for selected shift window","","",""),
        ("TC-R08","API","GET /api/reports/shift","Flask running","GET with params","GET /api/reports/shift?date=2026-05-19&shift=morning&plant=Plant001&area=AreaA","HTTP 200, JSON shift report","","",""),
        ("TC-R09","API","GET /api/reports/shifts","Flask running","GET list of shifts","GET /api/reports/shifts","HTTP 200, array of shift definitions {id, name, start_time, end_time}","","",""),
        ("TC-R10","Reports","Export shift report","Shift report loaded","Click Export","N/A","Excel file downloaded","","",""),
    ]
    for i, t in enumerate(tests2):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "MONTHLY REPORT", len(COLS))
    tests3 = [
        ("TC-R11","Reports","Monthly report page loads","Flask running","Navigate to Reports > Monthly","N/A","Month picker + plant/area dropdowns visible","","",""),
        ("TC-R12","Reports","Generate monthly report","Monthly data in DB","1. Select May 2026\n2. Click Generate","Month: 2026-05\nPlant: Plant001\nArea: AreaA","Report renders with daily averages for each day of the month","","",""),
        ("TC-R13","API","GET /api/reports/monthly","Flask running","GET with params","GET /api/reports/monthly?year=2026&month=5&plant=Plant001&area=AreaA","HTTP 200, JSON monthly report","","",""),
        ("TC-R14","Reports","No data month","Month with no historian data","Select month with no data\n(e.g. Jan 2020)","Month: 2020-01","Shows 'No data available' — not crash","","",""),
        ("TC-R15","Reports","Export monthly report","Monthly report loaded","Click Export Excel","N/A","Excel file downloaded","","",""),
    ]
    for i, t in enumerate(tests3):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 6. ADMIN ──────────────────────────────────────────────────
    ws = wb.create_sheet("6_Admin")
    row = write_sheet_header(ws, "Module 6 — Admin Panel (Users / Roles / Permissions)", COLS, WIDTHS)

    row = write_section_label(ws, row, "USER MANAGEMENT", len(COLS))
    tests = [
        ("TC-AD01","Admin","Admin page accessible","Logged in as Admin role","Navigate to Admin panel","N/A","Admin panel loads with tabs: Users, Roles, Permissions, Alerts","","",""),
        ("TC-AD02","Admin","Non-admin blocked","Logged in as non-admin user","Navigate to /admin","N/A","Redirect to dashboard with 'Access Denied' message","","",""),
        ("TC-AD03","Admin","List users","Admin logged in","1. Click Users tab","N/A","Table shows all users: username, email, role, status, last login","","",""),
        ("TC-AD04","Admin","Create new user","Admin logged in","1. Click Add User\n2. Fill form\n3. Submit","Username: testuser01\nEmail: test@test.com\nRole: Operator\nPassword: Test@1234","New user appears in list. Can login with those credentials.","","",""),
        ("TC-AD05","Admin","Disable user","User exists","1. Click Disable on user","User: testuser01","User status = disabled. Login attempt returns 'Account disabled'.","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "ROLES & PERMISSIONS", len(COLS))
    tests2 = [
        ("TC-AD06","Admin","List roles","Admin logged in","1. Click Roles tab","N/A","Table shows all roles with name and permission count","","",""),
        ("TC-AD07","Admin","Assign role to user","User + role exist","1. Edit user\n2. Change role","User: testuser01\nNew Role: Engineer","User role updated. New permissions apply on next login.","","",""),
        ("TC-AD08","Admin","Tag permissions — restrict","Admin logged in","1. Permissions tab\n2. Remove tag for role","Tag: Random.Real4\nRole: Operator","Operator users cannot see that tag in dashboard","","",""),
    ]
    for i, t in enumerate(tests2):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 7. ASSET BROWSER ──────────────────────────────────────────
    ws = wb.create_sheet("7_AssetBrowser")
    row = write_sheet_header(ws, "Module 7 — Asset Browser & Hierarchy", COLS, WIDTHS)

    row = write_section_label(ws, row, "ASSET HIERARCHY", len(COLS))
    tests = [
        ("TC-AS01","Assets","Asset browser loads","Hierarchy seeded in DB","Navigate to Asset Browser","N/A","Tree structure shows: Plant > Area > Equipment > Sub-equipment","","",""),
        ("TC-AS02","Assets","Click plant node","Hierarchy loaded","1. Click Plant001","N/A","Expands to show areas under Plant001","","",""),
        ("TC-AS03","Assets","Click equipment — view tags","Equipment has tags","1. Click specific equipment\n2. View tag panel","Equipment: Turbine001","Right panel shows live tag values for that equipment","","",""),
        ("TC-AS04","API","GET /api/equipment/hierarchy","Flask running","GET request","GET http://localhost:6001/api/equipment/hierarchy","HTTP 200, nested JSON {plant:[{area:[{equipment:[...]}]}]}","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 8. AUDIT TRAIL ────────────────────────────────────────────
    ws = wb.create_sheet("8_AuditTrail")
    row = write_sheet_header(ws, "Module 8 — Audit Trail", COLS, WIDTHS)

    row = write_section_label(ws, row, "AUDIT LOG", len(COLS))
    tests = [
        ("TC-AU01","Audit","Audit tab loads","Actions logged in DB","Navigate to Audit section","N/A","Table of actions: timestamp, user, action, target, IP","","",""),
        ("TC-AU02","Audit","Login creates audit entry","Audit enabled","1. Login\n2. Check audit log","User: Mustafa","Entry created: action='LOGIN', user='Mustafa', timestamp=now","","",""),
        ("TC-AU03","Audit","Alarm ACK creates audit","Alarm ACKed","1. ACK an alarm\n2. Check audit log","Alarm ID: any","Entry: action='ALARM_ACK', operator=user, alarm_id=N","","",""),
        ("TC-AU04","Audit","Filter audit by user","Audit log has entries","1. Filter by username\n2. Apply","Filter: user='Mustafa'","Only Mustafa's actions shown","","",""),
        ("TC-AU05","API","GET /api/audit","Flask running","GET with filter","GET /api/audit?user=Mustafa&limit=50","HTTP 200, array of audit events","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── FREEZE PANES + AUTOFILTER on all data sheets ──────────────
    for sh in wb.sheetnames:
        if sh != "COVER":
            wb[sh].freeze_panes = "A3"
            last_col = get_column_letter(10)
            wb[sh].auto_filter.ref = f"A2:{last_col}2"

    wb.save("TEST_FUNCTIONS_HMI.xlsx")
    print("  Saved: TEST_FUNCTIONS_HMI.xlsx")


# ═══════════════════════════════════════════════════════════════════
#  SERVER WORKBOOK
# ═══════════════════════════════════════════════════════════════════

def build_server_workbook():
    wb = openpyxl.Workbook()

    COLS = ["TC#", "Module", "Test Case Name", "Pre-Condition",
            "Test Steps", "Test Input / Values", "Expected Result", "Actual Result", "Status", "Notes"]
    WIDTHS = [6, 18, 28, 25, 40, 32, 35, 25, 10, 20]

    # ── COVER ────────────────────────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "COVER"
    write_cover(ws_cover,
        "Server-Side — Test Function Document",
        "Cereveate OPC DA Historian Platform  |  May 2026",
        [
            "Scope: C# OPC Backend (port 5001) + PostgreSQL/TimescaleDB + PLC Drivers",
            "",
            "Modules Covered:",
            "  1. OPC DA Connection & Tag Acquisition",
            "  2. Historian Ingest Pipeline (C# → PostgreSQL)",
            "  3. TimescaleDB — Hypertable, Compression, Policies",
            "  4. Rate Controller (Deadband / Change Detection)",
            "  5. PLC Drivers (Rockwell, Siemens, Modbus, Omron)",
            "  6. SignalR Real-Time Broadcast",
            "  7. Parquet Data Logging",
            "  8. Database Backup & Restore",
            "",
            "Test Environment:",
            "  C# Backend  :  http://localhost:5001",
            "  OPC Server  :  Matrikon.OPC.Simulation.1",
            "  DB          :  Automation_DB @ localhost:5432",
            "  PG Version  :  17.6  |  TimescaleDB: 2.23.0",
            "",
            "Legend:",
            "  GREEN row  = Expected PASS result",
            "  ORANGE row = Expected edge case / known behaviour",
            "  Status column: PASS / FAIL / SKIP",
        ]
    )

    # ── 1. OPC DA ─────────────────────────────────────────────────
    ws = wb.create_sheet("1_OPC_DA_Connection")
    row = write_sheet_header(ws, "Module 1 — OPC DA Connection & Tag Acquisition", COLS, WIDTHS)

    row = write_section_label(ws, row, "OPC SERVER CONNECT", len(COLS))
    tests = [
        ("TC-OPC01","OPC","Connect to local OPC server","Matrikon OPC running","1. Open HMI\n2. Click Connect on OPC panel\n3. Enter server prog ID","ProgID: Matrikon.OPC.Simulation.1\nHost: localhost","Connection status = Connected. Tags start updating.","","",""),
        ("TC-OPC02","OPC","Connect to remote OPC server","Remote PC with OPC + DCOM configured","Enter remote host","ProgID: Matrikon.OPC.Simulation.1\nHost: 192.168.1.XX","Connection established via DCOM. Tags flowing.","","","Requires DCOM config on both machines"),
        ("TC-OPC03","OPC","Invalid ProgID","OPC backend running","Enter wrong ProgID","ProgID: Invalid.Server.1\nHost: localhost","Connection fails with clear error. No crash.","","",""),
        ("TC-OPC04","OPC","OPC server offline","Matrikon not running","Try to connect","ProgID: Matrikon.OPC.Simulation.1","Connection fails, error logged. Backend stays running.","","",""),
        ("TC-OPC05","API","GET /api/opc/connections","C# backend running","GET request","GET http://localhost:5001/api/opc/connections","HTTP 200, array of connection objects with status","","",""),
        ("TC-OPC06","API","POST /api/opc/connect","C# backend running","POST to connect","Body: {progId:'Matrikon.OPC.Simulation.1', host:'localhost'}","HTTP 200, connection initiated","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "TAG READING & QUALITY", len(COLS))
    tests2 = [
        ("TC-OPC07","OPC","Read numeric tag","OPC connected","1. Check /api/opc/values\n2. Find a numeric tag","Tag: Random.Real4","value is float, quality='Good', timestamp recent","","",""),
        ("TC-OPC08","OPC","Read integer tag","OPC connected","Check tag value","Tag: Random.Int2","value is integer, quality='Good'","","",""),
        ("TC-OPC09","OPC","Read boolean tag","OPC connected","Check tag value","Tag: Random.Boolean","value is true/false, quality='Good'","","",""),
        ("TC-OPC10","OPC","Tag quality = Bad","OPC connected, tag removed","Check a removed/invalid tag","Tag: (non-existent tag ID)","quality='Bad', value may be null — NOT fake value generated","","","CRITICAL: no simulation"),
        ("TC-OPC11","OPC","Polling interval","OPC connected","1. Read tag at T=0\n2. Read again at T=1s\n3. Compare timestamps","Tag: Random.Real4","Timestamps differ by ~1000ms (OpcPollingIntervalMs=1000)","","",""),
        ("TC-OPC12","OPC","Tags refresh after reconnect","Connection dropped + restored","1. Disconnect OPC\n2. Reconnect\n3. Check values","N/A","Tags resume updating within 5s of reconnect","","",""),
    ]
    for i, t in enumerate(tests2):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 2. HISTORIAN INGEST ───────────────────────────────────────
    ws = wb.create_sheet("2_Historian_Ingest")
    row = write_sheet_header(ws, "Module 2 — Historian Ingest Pipeline", COLS, WIDTHS)

    row = write_section_label(ws, row, "TAG MASTER CONFIGURATION", len(COLS))
    tests = [
        ("TC-HI01","Historian","Tag master — insert enabled tag","DB accessible","Run SQL insert","INSERT INTO historian_meta.tag_master (tag_id, tag_name, data_type, deadband_value, db_logging_interval_ms, enabled) VALUES ('Random.Real4','Test Tag','double',0.5,1000,true)","Row inserted. No error.","","",""),
        ("TC-HI02","Historian","Tag master — disabled tag not written","DB accessible","1. Insert tag with enabled=false\n2. Wait 30s\n3. Check historian_timeseries","tag_id='Random.Real4', enabled=false","No new rows in historian_timeseries for that tag","","",""),
        ("TC-HI03","Historian","Tag master — upsert on conflict","Tag already exists","Run INSERT ON CONFLICT DO UPDATE","Same tag_id, enabled=true","Row updated, no duplicate error","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "DATA WRITE VERIFICATION", len(COLS))
    tests2 = [
        ("TC-HI04","Historian","Data written to DB","tag_master enabled, OPC connected","1. Enable tag in tag_master\n2. Wait 10s\n3. Query","SELECT count(*) FROM historian_raw.historian_timeseries WHERE tag_id='Random.Real4' AND time > now()-interval '1 min'","Returns ≥ 1 row (new data written)","","",""),
        ("TC-HI05","Historian","No duplicate writes","tag_master enabled","1. Enable tag\n2. Wait 60s\n3. Check for dupes","SELECT tag_id, time, count(*) FROM historian_raw.historian_timeseries WHERE tag_id='Random.Real4' GROUP BY tag_id,time HAVING count(*)>1","0 rows — no duplicates","","",""),
        ("TC-HI06","Historian","Timestamp is OPC timestamp","OPC connected","Query latest row","SELECT tag_id, time, value FROM historian_raw.historian_timeseries WHERE tag_id='Random.Real4' ORDER BY time DESC LIMIT 1","time matches OPC server timestamp (not server insert time)","","",""),
        ("TC-HI07","Historian","Quality stored correctly","OPC connected","Query quality column","SELECT quality FROM historian_raw.historian_timeseries WHERE tag_id='Random.Real4' ORDER BY time DESC LIMIT 5","Quality = 'G' (Good) for active OPC tags","","",""),
    ]
    for i, t in enumerate(tests2):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 3. RATE CONTROLLER ────────────────────────────────────────
    ws = wb.create_sheet("3_RateController")
    row = write_sheet_header(ws, "Module 3 — Rate Controller (Deadband / Change Detection)", COLS, WIDTHS)

    row = write_section_label(ws, row, "DEADBAND LOGIC", len(COLS))
    tests = [
        ("TC-RC01","RateCtrl","Deadband=0: write on any change","Tag with deadband_value=0","1. Enable tag\n2. Wait for value change\n3. Check DB","deadband_value=0\nValue changes: 10.0 → 10.1","New row written for every change (exact comparison)","","",""),
        ("TC-RC02","RateCtrl","Deadband>0: change < threshold = NO write","Tag with deadband_value=1.0","1. Tag value changes by 0.5\n2. Check DB","deadband=1.0\nChange: 10.0 → 10.4 (delta=0.4 < 1.0)","No new row written for this change","","",""),
        ("TC-RC03","RateCtrl","Deadband>0: change > threshold = write","Tag with deadband_value=1.0","1. Tag value changes by 2.0\n2. Check DB","deadband=1.0\nChange: 10.0 → 12.1 (delta=2.1 > 1.0)","New row written immediately (spike detection)","","",""),
        ("TC-RC04","RateCtrl","First sample always written","New tag added","1. Add tag to tag_master\n2. Wait for first poll","Any tag_id","First ever row for that tag is always written regardless of deadband","","",""),
        ("TC-RC05","RateCtrl","Interval check","Tag with 5000ms interval","1. Enable tag\n2. Monitor writes over 10s","db_logging_interval_ms=5000","Writes occur at most once per 5 seconds","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 4. TIMESCALEDB ────────────────────────────────────────────
    ws = wb.create_sheet("4_TimescaleDB")
    row = write_sheet_header(ws, "Module 4 — TimescaleDB Hypertable, Compression & Policies", COLS, WIDTHS)

    row = write_section_label(ws, row, "HYPERTABLE VERIFICATION", len(COLS))
    tests = [
        ("TC-TS01","TimescaleDB","historian_timeseries is a hypertable","DB running","Run SQL","SELECT hypertable_name FROM timescaledb_information.hypertables WHERE hypertable_schema='historian_raw'","Returns 'historian_timeseries'","","",""),
        ("TC-TS02","TimescaleDB","Chunk count","Data in DB","Run SQL","SELECT count(*) FROM timescaledb_information.chunks WHERE hypertable_name='historian_timeseries'","Returns ≥ 1 chunk","","",""),
        ("TC-TS03","TimescaleDB","Compression stats","Compression ran","Run SQL","SELECT pg_size_pretty(before_compression_total_bytes) AS before, pg_size_pretty(after_compression_total_bytes) AS after FROM chunk_compression_stats('historian_raw.historian_timeseries') LIMIT 3","before > after — shows compression working. Ratio ~19x expected.","","",""),
        ("TC-TS04","TimescaleDB","TimescaleDB extension version","DB running","Run SQL","SELECT extversion FROM pg_extension WHERE extname='timescaledb'","Returns '2.23.0' or newer","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "POLICIES & BACKGROUND JOBS", len(COLS))
    tests2 = [
        ("TC-TS05","TimescaleDB","Compression policy exists","DB running","Run SQL","SELECT * FROM timescaledb_information.jobs WHERE application_name LIKE '%Compress%' AND hypertable_name='historian_timeseries'","1 row returned — policy job active","","",""),
        ("TC-TS06","TimescaleDB","Retention policy exists","DB running","Run SQL","SELECT * FROM timescaledb_information.jobs WHERE application_name LIKE '%Retention%' AND hypertable_name='historian_timeseries'","1 row returned — drop after 2 years policy active","","",""),
        ("TC-TS07","TimescaleDB","Continuous aggregate job","DB running","Run SQL","SELECT * FROM timescaledb_information.jobs WHERE application_name LIKE '%Refresh%'","1 row — ts_hourly_agg refresh job, runs every 1h","","",""),
        ("TC-TS08","TimescaleDB","Last job run success","DB running","Run SQL","SELECT application_name, last_run_status, last_successful_finish FROM timescaledb_information.job_stats JOIN timescaledb_information.jobs USING (job_id) WHERE hypertable_schema='historian_raw'","last_run_status = 'Success' for all 3 jobs","","",""),
        ("TC-TS09","TimescaleDB","BRIN index exists","DB running","Run SQL","SELECT indexname FROM pg_indexes WHERE tablename='historian_timeseries' AND indexdef LIKE '%brin%'","Returns 'idx_historian_ts_time_brin'","","",""),
        ("TC-TS10","TimescaleDB","B-tree tag+time index exists","DB running","Run SQL","SELECT indexname FROM pg_indexes WHERE tablename='_hyper%' AND indexdef LIKE '%tag_id%'","Returns idx_historian_ts_tagid_time or similar","","",""),
    ]
    for i, t in enumerate(tests2):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "CONTINUOUS AGGREGATE", len(COLS))
    tests3 = [
        ("TC-TS11","TimescaleDB","ts_hourly_agg has data","Data in historian","Run SQL","SELECT count(*) FROM historian_raw.ts_hourly_agg","Returns ≥ 6000","","",""),
        ("TC-TS12","TimescaleDB","Hourly agg columns correct","ts_hourly_agg exists","Run SQL","SELECT bucket, tag_id, avg_value, min_value, max_value, sample_count FROM historian_raw.ts_hourly_agg LIMIT 1","All columns return values — no null in bucket/tag_id","","",""),
        ("TC-TS13","TimescaleDB","Agg refreshes on new data","Data flowing","1. Wait 1h (or manually refresh)\n2. Check max bucket","CALL refresh_continuous_aggregate('historian_raw.ts_hourly_agg', now()-interval '2h', now())","New hourly buckets appear for recent data","","",""),
    ]
    for i, t in enumerate(tests3):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 5. PLC DRIVERS ────────────────────────────────────────────
    ws = wb.create_sheet("5_PLC_Drivers")
    row = write_sheet_header(ws, "Module 5 — PLC Driver Tests", COLS, WIDTHS)

    row = write_section_label(ws, row, "ROCKWELL (Allen-Bradley) — libplctag", len(COLS))
    tests = [
        ("TC-PLC01","Rockwell","Connect to Rockwell PLC","PLC online, IP reachable","1. Configure PlcProtocol.Rockwell\n2. Set IP + tag path\n3. Read tag","IP: <PLC IP>\nTag: B3:0/0 or INT[0]","Tag value returned, quality=Good","","","Driver: RockwellDriver.cs"),
        ("TC-PLC02","Rockwell","PLC offline — no fake values","PLC unreachable","Attempt to read when PLC offline","IP: unreachable","Returns null with quality=Bad or CommError — NOT fake/random value","","","CRITICAL: no simulation"),
        ("TC-PLC03","Rockwell","Wrong tag path","PLC online","Read non-existent tag","Tag: NonExistent_Tag","Returns error/null — no crash","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "SIEMENS S7 — S7.Net", len(COLS))
    tests2 = [
        ("TC-PLC04","SiemensS7","Connect to S7 PLC","S7 PLC online","Configure PlcProtocol.SiemensS7, IP, rack/slot","IP: <PLC IP>\nRack:0, Slot:1","Connection established, tags readable","","","Driver: SiemensS7Driver.cs"),
        ("TC-PLC05","SiemensS7","Read DB block","S7 connected","Read data block value","Address: DB1.DBD0 (REAL)","Float value returned","","",""),
        ("TC-PLC06","SiemensS7","PLC offline — CommError","S7 PLC unreachable","Read when offline","N/A","Quality=CommError, value=null — no fake value","","","CRITICAL: no simulation"),
    ]
    for i, t in enumerate(tests2):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "MODBUS TCP — NModbus", len(COLS))
    tests3 = [
        ("TC-PLC07","ModbusTcp","Connect Modbus device","Modbus slave online","Configure PlcProtocol.ModbusTcp, IP, port 502","IP: <device IP>\nPort: 502","Connection established","","","Driver: ModbusTcpDriver.cs"),
        ("TC-PLC08","ModbusTcp","Read holding register","Modbus connected","Read register","Register: 40001 (HR[0])","Integer value returned, quality=Good","","",""),
        ("TC-PLC09","ModbusTcp","Read coil","Modbus connected","Read coil","Coil: 1 (C[0])","Boolean value returned","","",""),
        ("TC-PLC10","ModbusTcp","Device offline — no fake","Modbus device unreachable","Attempt read when offline","N/A","null value, quality=Bad — NOT fake","","","CRITICAL: no simulation"),
    ]
    for i, t in enumerate(tests3):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "DELETED DRIVER — EtherNet/IP", len(COLS))
    tests4 = [
        ("TC-PLC11","EtherNetIP","EtherNetIP throws error","C# backend running","Configure PlcProtocol.EtherNetIP, attempt read","Any tag","NotImplementedException thrown — clear error message, no crash, no fake values","","","Driver deleted Dec 2025. Use Rockwell for AB PLCs."),
    ]
    for i, t in enumerate(tests4):
        row = write_data_row(ws, row, t, alt=False, highlight="warn")

    # ── 6. SIGNALR ────────────────────────────────────────────────
    ws = wb.create_sheet("6_SignalR")
    row = write_sheet_header(ws, "Module 6 — SignalR Real-Time Broadcast", COLS, WIDTHS)

    row = write_section_label(ws, row, "SIGNALR HUB", len(COLS))
    tests = [
        ("TC-SR01","SignalR","Hub endpoint accessible","C# backend running","1. Open browser DevTools > Network\n2. Connect to HMI\n3. Check WS connections","ws://localhost:5001/opcHub","WebSocket upgrade successful, 101 Switching Protocols","","",""),
        ("TC-SR02","SignalR","TagValuesUpdated event fires","OPC connected","1. Monitor WS frames in DevTools\n2. Wait 2s","N/A","TagValuesUpdated message arrives every ~1000ms","","",""),
        ("TC-SR03","SignalR","Multiple clients get updates","2 browser tabs open","1. Open tab 1 on dashboard\n2. Open tab 2 on dashboard\n3. Observe both","N/A","Both tabs receive live updates simultaneously","","",""),
        ("TC-SR04","SignalR","Reconnect on disconnect","Connected, then network drop","1. Disable network briefly\n2. Re-enable\n3. Check updates resume","N/A","SignalR auto-reconnects, tag updates resume within 10s","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 7. PARQUET LOGGING ────────────────────────────────────────
    ws = wb.create_sheet("7_Parquet_Logging")
    row = write_sheet_header(ws, "Module 7 — Parquet Data Logging", COLS, WIDTHS)

    row = write_section_label(ws, row, "PARQUET FILE CREATION", len(COLS))
    tests = [
        ("TC-PQ01","Parquet","Parquet files created","OPC connected, SelectedTags in logging-config.json","1. Wait 5s after OPC connect\n2. Check D:\\OpcLogs\\Data\\","N/A","New .parquet file appears — not empty","","",""),
        ("TC-PQ02","Parquet","Only SelectedTags logged","logging-config.json configured","1. Check parquet column names\n2. Compare to SelectedTags list","SelectedTags: [Random.Real4, ...]","Only configured tags appear as columns — no extra tags","","",""),
        ("TC-PQ03","Parquet","File rotation at 10MB","Large data volume","1. Wait for file size to reach 10MB\n2. Check folder","N/A","New .parquet file created, old one closed. Both readable.","","",""),
        ("TC-PQ04","Parquet","Polling interval = 1000ms","OPC connected","1. Read parquet timestamps\n2. Check row interval","N/A","Rows spaced ~1000ms apart (OpcPollingIntervalMs=1000)","","",""),
        ("TC-PQ05","Parquet","No null rows from bad quality","OPC tag with Bad quality","1. Disconnect one OPC tag\n2. Check parquet","Tag with Bad quality","Row written with null value — NOT fake value","","","CRITICAL: no simulation"),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # ── 8. BACKUP & RESTORE ───────────────────────────────────────
    ws = wb.create_sheet("8_Backup_Restore")
    row = write_sheet_header(ws, "Module 8 — Database Backup & Restore", COLS, WIDTHS)

    row = write_section_label(ws, row, "BACKUP", len(COLS))
    tests = [
        ("TC-BK01","Backup","Backup script runs successfully","DB running, pg_dump available","Run: python backup_automation_db.py","N/A","C:\\DB_Backups\\ contains 3 files: .dump, _roles.sql, _info.txt","","",""),
        ("TC-BK02","Backup","Dump file size reasonable","Backup created","Check file size","N/A","Dump file ≥ 50 MB (currently ~118 MB). Not 0 bytes.","","",""),
        ("TC-BK03","Backup","Roles file not empty","Backup created","Check roles SQL file","N/A","File contains CREATE ROLE / CREATE USER statements","","",""),
        ("TC-BK04","Backup","Manifest file accurate","Backup created","Open info.txt, check contents","N/A","Shows correct date, size, schema list","","",""),
    ]
    for i, t in enumerate(tests):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    row = write_section_label(ws, row, "RESTORE VERIFICATION (Post-Restore Checks)", len(COLS))
    tests2 = [
        ("TC-BK05","Restore","TimescaleDB extension present after restore","Restored DB","Run SQL","SELECT extversion FROM pg_extension WHERE extname='timescaledb'","Returns '2.23.0'","","",""),
        ("TC-BK06","Restore","Row count matches source","Restored DB","Run SQL","SELECT count(*) FROM historian_raw.historian_timeseries","Matches source count (≥15,000,000)","","",""),
        ("TC-BK07","Restore","All chunks present","Restored DB","Run SQL","SELECT count(*) FROM timescaledb_information.chunks WHERE hypertable_name='historian_timeseries'","Same chunk count as source","","",""),
        ("TC-BK08","Restore","Compression policies active","Restored DB","Run SQL","SELECT count(*) FROM timescaledb_information.jobs WHERE hypertable_name='historian_timeseries'","3 jobs (compress, retain, agg refresh)","","",""),
        ("TC-BK09","Restore","restoring flag REMOVED","After restore complete","Check postgresql.conf","grep timescaledb.restoring postgresql.conf","Line must NOT be present or must be commented out","","","CRITICAL — jobs won't run if flag left on"),
        ("TC-BK10","Restore","New data writes after restore","Restored, OPC connected","1. Connect OPC\n2. Enable tag in tag_master\n3. Wait 30s","N/A","New rows appear in historian_timeseries — restore fully operational","","",""),
    ]
    for i, t in enumerate(tests2):
        row = write_data_row(ws, row, t, alt=(i%2==0))

    # freeze + autofilter
    for sh in wb.sheetnames:
        if sh != "COVER":
            wb[sh].freeze_panes = "A3"
            last_col = get_column_letter(10)
            wb[sh].auto_filter.ref = f"A2:{last_col}2"

    wb.save("TEST_FUNCTIONS_SERVER.xlsx")
    print("  Saved: TEST_FUNCTIONS_SERVER.xlsx")


# ─── MAIN ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n  Generating Test Function Documents...")
    build_hmi_workbook()
    build_server_workbook()
    print("\n  Done. Files saved in project root:")
    print("    TEST_FUNCTIONS_HMI.xlsx")
    print("    TEST_FUNCTIONS_SERVER.xlsx\n")
