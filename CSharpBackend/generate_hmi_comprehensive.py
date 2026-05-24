"""
Generates a SINGLE comprehensive HMI test function Excel:
  TEST_FUNCTIONS_HMI_COMPREHENSIVE.xlsx

Modules:
  COVER, Authentication, Dashboard/Live, Alarms, Trends/Historical,
  Reports-Daily, Reports-Shift, Reports-Monthly,
  DATA MATCHING (report calc verification),
  DB Logs Verification, Admin, Asset Browser, Audit Trail
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── COLOURS ─────────────────────────────────────────────────────
C_TITLE      = "1F3864"
C_SEC_BLUE   = "2E75B6"
C_COL_HDR    = "D6E4F0"
C_PASS       = "C6EFCE"
C_WARN       = "FCE4D6"
C_CRITICAL   = "FF0000"
C_ALT        = "F2F7FB"
C_WHITE      = "FFFFFF"
C_GOLD       = "FFC000"
C_PURPLE     = "7030A0"
C_PURPLE_LT  = "EAD6F5"
C_GREEN_DARK = "375623"
C_ORANGE     = "F4B942"

def fill(c): return PatternFill("solid", fgColor=c)
def bdr():
    s = Side(style="thin", color="B0B8C1")
    return Border(left=s, right=s, top=s, bottom=s)
def hfont(sz=10, bold=True, color="FFFFFF"):
    return Font(name="Calibri", size=sz, bold=bold, color=color)
def bfont(sz=9, bold=False, color="1F3864"):
    return Font(name="Calibri", size=sz, bold=bold, color=color)
def walign(): return Alignment(wrap_text=True, vertical="center", horizontal="left")
def calign(): return Alignment(wrap_text=True, vertical="center", horizontal="center")

# ─── SHEET COLUMNS ───────────────────────────────────────────────
COLS   = ["TC#","Module","Test Case Name","Pre-Condition",
          "Test Steps","Test Input / Values","Expected Result",
          "Actual Result","Status","Remarks / Notes"]
WIDTHS = [6, 16, 30, 26, 42, 34, 38, 26, 9, 22]

def setup_sheet(ws, title):
    ws.sheet_view.showGridLines = False
    span = len(COLS)
    last = get_column_letter(span)

    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value = title
    c.font  = hfont(13)
    c.fill  = fill(C_TITLE)
    c.alignment = calign()
    ws.row_dimensions[1].height = 32

    for i,(h,w) in enumerate(zip(COLS,WIDTHS),1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = w
        cell = ws.cell(row=2,column=i,value=h)
        cell.font      = Font(name="Calibri",size=10,bold=True,color=C_TITLE)
        cell.fill      = fill(C_COL_HDR)
        cell.alignment = calign()
        cell.border    = bdr()
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last}2"
    return 3

def sec(ws, row, label, colour=C_GOLD):
    span = len(COLS)
    last = get_column_letter(span)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws[f"A{row}"]
    c.value     = f"  ▶  {label}"
    c.font      = Font(name="Calibri",size=10,bold=True,color=C_TITLE)
    c.fill      = fill(colour)
    c.alignment = walign()
    ws.row_dimensions[row].height = 20
    return row+1

def row_write(ws, row, vals, alt=False, hi=None):
    bg = C_ALT if alt else C_WHITE
    if hi == "pass":   bg = C_PASS
    elif hi == "warn": bg = C_WARN
    elif hi == "crit": bg = "FFCCCC"
    elif hi == "calc": bg = C_PURPLE_LT

    for col,v in enumerate(vals,1):
        cell = ws.cell(row=row,column=col,value=v)
        cell.fill      = fill(bg)
        cell.font      = bfont()
        cell.border    = bdr()
        cell.alignment = walign()
    ws.row_dimensions[row].height = 46
    return row+1

# ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════
# COVER
# ═══════════════════════════════════════════════════════════════
wc = wb.active
wc.title = "COVER"
wc.sheet_view.showGridLines = False
wc.column_dimensions["A"].width = 4
wc.column_dimensions["B"].width = 60
wc.column_dimensions["C"].width = 28

wc.merge_cells("B2:C2")
t=wc["B2"]; t.value="HMI — Comprehensive Test Function Document"
t.font=Font(name="Calibri",size=20,bold=True,color="FFFFFF")
t.fill=fill(C_TITLE); t.alignment=walign(); wc.row_dimensions[2].height=50

wc.merge_cells("B3:C3")
t=wc["B3"]; t.value="Cereveate OPC DA Historian Platform  |  May 2026"
t.font=Font(name="Calibri",size=12,bold=False,color="FFFFFF")
t.fill=fill(C_SEC_BLUE); t.alignment=walign(); wc.row_dimensions[3].height=26

lines=[
    "","Scope: React/Vite HMI (port 8090)  +  Flask API (port 6001)","",
    "Modules:","  1.  Authentication & MFA",
    "  2.  Dashboard & Live Tag Values",
    "  3.  Alarm Management (Active / ACK / History / Suppress / Trips)",
    "  4.  Trend & Historical Data (chart, downsample, export)",
    "  5.  Report — Daily  (hourly avg/max/min, 5 AM–5 AM cycle)",
    "  6.  Report — Shift  (Morning 06-14 / Evening 14-22 / Night 22-06)",
    "  7.  Report — Monthly (daily rollup averages)",
    "  8.  DATA MATCHING — Manual calc verification vs DB values  ← NEW",
    "  9.  DB Logs Verification (historian writes, agg rows, audit)  ← NEW",
    "  10. Admin Panel (Users / Roles / Permissions / Alerts)",
    "  11. Asset Browser & Equipment Hierarchy",
    "  12. Audit Trail",
    "",
    "Test Environment:",
    "  HMI URL        :  http://localhost:8090",
    "  API Base       :  http://localhost:6001",
    "  Login          :  Mustafa  /  Admin@123",
    "  DB             :  Automation_DB  @  localhost:5432",
    "  Timezone       :  Asia/Kolkata  (IST = UTC+5:30)",
    "  Day boundary   :  Reports run 05:00 IST → 04:59 IST next day",
    "",
    "Highlights of this version:",
    "  • All original TCs from previous document included",
    "  • Module 8 — full manual calculation cross-check (avg, max, min)",
    "  • Module 9 — DB log verification with exact SQL checks",
    "  • PURPLE rows = calculation verification steps (do manually)",
    "  • GREEN rows  = expected PASS result",
    "  • ORANGE rows = edge case / known boundary behaviour",
]
for i,l in enumerate(lines,5):
    wc.row_dimensions[i].height=16
    wc.merge_cells(f"B{i}:C{i}")
    c=wc[f"B{i}"]; c.value=l
    c.font=Font(name="Calibri",size=10); c.alignment=walign()

# ═══════════════════════════════════════════════════════════════
# 1 — AUTH
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("1_Authentication")
row=setup_sheet(ws,"Module 1 — Authentication & MFA")
row=sec(ws,row,"LOGIN")
data=[
("TC-A01","Auth","Valid login — Admin","App on port 8090","1. Go to http://localhost:8090\n2. Enter credentials\n3. Click Login","Username: Mustafa\nPassword: Admin@123","Redirect to MFA screen or dashboard. HTTP 200 from /api/auth/login.","","",""),
("TC-A02","Auth","Wrong password","App running","Enter correct user, wrong password","Username: Mustafa\nPassword: WrongPass1!","Error shown: 'Invalid credentials'. No redirect. HTTP 401.","","",""),
("TC-A03","Auth","Non-existent user","App running","Enter unknown username","Username: ghost_user99\nPassword: anything","Error: 'Invalid credentials'. HTTP 401.","","",""),
("TC-A04","Auth","Empty username field","App running","Leave username blank, click Login","Username: (blank)\nPassword: Admin@123","Field validation error — 'Username required'. No API call.","","",""),
("TC-A05","Auth","Empty password field","App running","Leave password blank, click Login","Username: Mustafa\nPassword: (blank)","Field validation error — 'Password required'.","","",""),
("TC-A06","Auth","SQL injection in username","App running","Enter SQL in username","Username: admin' OR '1'='1\nPassword: x","Login fails with 'Invalid credentials'. No DB error exposed.","","","Security check"),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"MFA (Multi-Factor Authentication)")
data=[
("TC-A07","MFA","Correct TOTP code","Login step passed, MFA enabled","Enter valid 6-digit code from authenticator app","TOTP: from Google Authenticator","Redirect to dashboard. Session token issued.","","",""),
("TC-A08","MFA","Wrong TOTP code","Login step passed","Enter wrong 6 digits","TOTP: 000000","Error: 'Invalid MFA code'. Stay on MFA screen.","","",""),
("TC-A09","MFA","Expired TOTP (>30s old)","Login step passed","Enter code > 30 seconds old","Old TOTP code","Error: 'Invalid or expired MFA code'.","","","TOTP window = 30s"),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"SESSION & LOGOUT")
data=[
("TC-A10","Auth","Logout clears session","Logged in","Click Logout","N/A","Redirect to /login. JWT gone from browser. Next API call returns 401.","","",""),
("TC-A11","Auth","Expired JWT rejected","Have expired token","Make API call with old JWT in header","Authorization: Bearer <expired_token>","HTTP 401. JSON: {error:'Token expired'}.","","",""),
("TC-A12","Auth","GET /api/auth/me with valid token","Logged in","GET /api/auth/me with Bearer token","Header: Authorization: Bearer <valid_token>","HTTP 200. Returns {user_id, username, role}.","","",""),
("TC-A13","Auth","GET /api/auth/me without token","No token","GET /api/auth/me no header","Header: (none)","HTTP 401 Unauthorized.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

# ═══════════════════════════════════════════════════════════════
# 2 — DASHBOARD
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("2_Dashboard_LiveTags")
row=setup_sheet(ws,"Module 2 — Dashboard & Live Tag Values")
row=sec(ws,row,"LIVE TAG DISPLAY")
data=[
("TC-D01","Dashboard","Tags load on page open","OPC connected, tags configured","1. Login\n2. Navigate to Dashboard","N/A","Tag table/cards visible. Values not blank.","","",""),
("TC-D02","Dashboard","Tag values update every ~1s","OPC connected","1. Observe any numeric tag (e.g. Random.Real4)\n2. Watch for 3s","Tag: Random.Real4","Value changes on screen within 2–3 seconds.","","","Polling = 1000ms"),
("TC-D03","Dashboard","Bad quality shown, not fake","OPC tag offline","1. Disconnect OPC server\n2. Watch tags","N/A","Quality indicator turns Bad/red. Values show last known or null. No random/fake values generated.","","","CRITICAL: no simulation"),
("TC-D04","Dashboard","Tag count matches config","logging-config.json known","Count tag cards on dashboard","N/A","Number of tags shown = number of tags in OPC connection","","",""),
("TC-D05","API","GET /api/opc/values — healthy","OPC connected","GET http://localhost:5001/api/opc/values","N/A","HTTP 200. JSON array. Each item has: tagId, value, quality='Good', timestamp.","","",""),
("TC-D06","API","GET /api/opc/values — OPC down","OPC disconnected","GET /api/opc/values","N/A","HTTP 200. quality='Bad' on all tags. No fake values.","","","CRITICAL"),
("TC-D07","SignalR","WS connection active","Browser open","Open DevTools > Network > WS tab","ws://localhost:5001/opcHub","WebSocket 101 upgrade visible. TagValuesUpdated frames arriving.","","",""),
("TC-D08","SignalR","Two clients both update","2 tabs open","Open dashboard in Tab1 and Tab2","N/A","Both tabs receive updates independently at ~1s interval.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

# ═══════════════════════════════════════════════════════════════
# 3 — ALARMS
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("3_Alarms")
row=setup_sheet(ws,"Module 3 — Alarm Management")
row=sec(ws,row,"ACTIVE ALARMS")
data=[
("TC-AL01","Alarms","Active alarm list loads","Flask running, alarms in DB","Navigate to Alarms page > Active tab","N/A","Table shows: Tag ID, Description, Severity, Timestamp, ACK status. No JS error.","","",""),
("TC-AL02","Alarms","Alarm count badge on nav","Active alarms exist","Check sidebar navigation","N/A","Alarm icon shows numeric badge = count from /api/alarms/stats.","","",""),
("TC-AL03","API","GET /api/alarms/active","Flask running","GET with auth header","GET http://localhost:6001/api/alarms/active\nHeader: Bearer <token>","HTTP 200. {success:true, alarms:[...], count:N}","","",""),
("TC-AL04","API","GET /api/alarms/stats","Flask running","GET request","GET /api/alarms/stats","HTTP 200. {critical:N, high:N, medium:N, low:N}","","",""),
("TC-AL05","Alarms","Filter alarms by severity","Active alarms of mixed severity","Click severity filter = CRITICAL","N/A","Only critical alarms shown in table.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"ACKNOWLEDGE & CLEAR")
data=[
("TC-AL06","Alarms","Acknowledge single alarm","Unacked active alarm","1. Click ACK on alarm row\n2. Enter comment","Alarm ID: from active list\nComment: 'Acked by Mustafa'","Alarm ack_status = true. Audit entry created. Green tick shown.","","",""),
("TC-AL07","API","POST /api/alarms/acknowledge/<id>","Active alarm in DB","POST with body","POST /api/alarms/acknowledge/1\nBody: {comment:'test ack'}","HTTP 200. {success:true}. DB updated.","","",""),
("TC-AL08","Alarms","Acknowledge by Tag ID","Tag with active alarm","POST acknowledge-by-tag","tag_id: Random.Real4\nComment: 'Bulk ack'","All alarms for that tag become acknowledged.","","",""),
("TC-AL09","Alarms","Clear an alarm","ACKed alarm","1. Click Clear button","ACKed alarm","Alarm moves to history. No longer in active list.","","",""),
("TC-AL10","API","POST /api/alarms/clear/<id>","ACKed alarm","POST clear","POST /api/alarms/clear/1","HTTP 200. Alarm removed from active table.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"ALARM HISTORY & AUDIT")
data=[
("TC-AL11","Alarms","History tab loads","Historical alarms exist","Navigate to Alarm History tab","N/A","Table: past alarms, timestamps, operator, action.","","",""),
("TC-AL12","API","GET /api/alarms/history with dates","Flask running","GET with date range","GET /api/alarms/history?start=2026-05-01&end=2026-05-20","HTTP 200. Array of events in range.","","",""),
("TC-AL13","Alarms","Audit trail per alarm","Alarm with history","Click audit icon on alarm","Any alarm ID","Popup shows: Created → Acknowledged → Cleared with operator+time each.","","",""),
("TC-AL14","API","GET /api/alarms/audit/<id>","Alarm with audit","GET request","GET /api/alarms/audit/1","HTTP 200. Array of audit steps.","","",""),
("TC-AL15","API","GET /api/alarms/audit/tag/<tag_id>","Tag with alarm history","GET by tag","GET /api/alarms/audit/tag/Random.Real4","HTTP 200. All audit events for that tag.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"SUPPRESSION & TRIPS")
data=[
("TC-AL16","Alarms","Suppress alarm","Active alarm","1. Click Suppress\n2. Set 1-hour window","Duration: 60 min\nReason: Maintenance","Alarm suppressed. Shown in Suppressed list. No new notification during window.","","",""),
("TC-AL17","API","POST /api/alarms/suppress/<id>","Active alarm","POST","POST /api/alarms/suppress/1\nBody:{duration_minutes:60,reason:'Maint'}","HTTP 200.","","",""),
("TC-AL18","API","GET /api/alarms/suppressed","Flask running","GET","GET /api/alarms/suppressed","HTTP 200. List of suppressed alarms.","","",""),
("TC-AL19","API","GET /api/alarms/trips","Flask running","GET","GET /api/alarms/trips","HTTP 200. Array of trip events.","","",""),
("TC-AL20","API","GET /api/alarms/interlocks","Flask running","GET","GET /api/alarms/interlocks","HTTP 200. Array of interlock states.","","",""),
("TC-AL21","Alarms","Unacknowledged summary","Unacked alarms","GET unacknowledged endpoint","GET /api/alarms/audit/unacknowledged","HTTP 200. List of alarms awaiting acknowledgement.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

# ═══════════════════════════════════════════════════════════════
# 4 — TRENDS
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("4_Trends_Historical")
row=setup_sheet(ws,"Module 4 — Trend & Historical Data")
row=sec(ws,row,"TREND CHART — UI")
data=[
("TC-TR01","Trends","Historian page loads","Historian DB populated","Navigate to Historian/Backup tab","N/A","Page loads. Tag selector and date picker visible.","","",""),
("TC-TR02","Trends","Single tag last 24h","DB has data for tag","1. Select Random.Real4\n2. Range: last 24h\n3. Load","Tag: Random.Real4\nRange: now-24h → now","Chart renders. X=time, Y=value. Points connected.","","",""),
("TC-TR03","Trends","Two tags same chart","2+ tags in DB","Select 2 tags, load","Tags: Random.Real4, Random.Int2","Two coloured lines. Legend shows both tag names.","","",""),
("TC-TR04","Trends","Empty date range","Period with no data","Select 2020-01-01 → 2020-01-02","Start: 2020-01-01\nEnd: 2020-01-02","'No data available' message. No crash.","","",""),
("TC-TR05","Trends","Export CSV","Data loaded in chart","Click Export CSV","N/A","File downloads. Columns: timestamp, tag_id, value.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"DOWNSAMPLING VERIFICATION (Manual Calc)")
data=[
("TC-TR06","Trends","Downsampling — total > 1000 pts","DB has >1000 rows for tag in range","1. Select tag with large range (e.g. 7 days)\n2. Count data points on chart","Tag: Random.Real4\nRange: 7 days","Chart returns ≤ 1000 points. Server applied epoch-bucket AVG downsampling.","","","Auto-downsamples when total > max_points"),
("TC-TR07","Trends-Calc","Verify downsampled AVG vs raw","DB accessible","1. Get raw rows from DB for one bucket:\nSELECT AVG(value_num) FROM historian_raw.historian_timeseries WHERE tag_id='Random.Real4' AND time BETWEEN '<bucket_start>' AND '<bucket_end>'\n2. Compare to chart value for that point","Expected: chart avg ≈ DB AVG (within 0.01)","Chart value matches manual AVG calculation from raw DB rows.","","","CALC VERIFY — purple"),
("TC-TR08","Trends-Calc","Sampling interval mode","DB has data","Select explicit 30-second interval\nCount returned points","Range: 1 hour, interval: 30s\nExpected points: 3600/30 = 120","API returns ≤ 120 points. Each timestamp is on a 30s boundary.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0,hi="calc")

row=sec(ws,row,"API CALLS")
data=[
("TC-TR09","API","GET /api/historical/trend — single tag","DB running","GET with params","GET /api/historical/trend?tag_id=Random.Real4&start=2026-05-19T00:00:00&end=2026-05-20T00:00:00","HTTP 200. {data:[{timestamp,value,quality},...]}","","",""),
("TC-TR10","API","GET /api/historical/trend — no token","Flask running","GET without auth","GET /api/historical/trend?tag_id=Random.Real4","HTTP 401 Unauthorized.","","",""),
("TC-TR11","API","GET /api/historical/tags","Flask running","GET tag list","GET /api/historical/tags","HTTP 200. Array of tag IDs that have historian data.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

# ═══════════════════════════════════════════════════════════════
# 5 — DAILY REPORT
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("5_Report_Daily")
row=setup_sheet(ws,"Module 5 — Daily Report (5 AM – 5 AM cycle, 24 hourly columns)")
row=sec(ws,row,"UI LOAD & GENERATION")
data=[
("TC-RD01","Daily Rpt","Daily report page loads","Flask running, templates seeded","Navigate to Reports > Daily","N/A","Date picker, plant/area dropdowns visible. No JS error.","","",""),
("TC-RD02","Daily Rpt","Generate report — valid date","Historian data for 2026-05-19","Select date 2026-05-19, Plant001, AreaA, click Generate","Date: 2026-05-19\nPlant: Plant001\nArea: AreaA","Table renders. 24 hourly columns (5am→4am). Rows = tags in template.","","",""),
("TC-RD03","Daily Rpt","Day boundary = 05:00 IST","Report for 2026-05-19","Check first and last hour column headers","N/A","First column = '5 am To 6 am'. Last = '4 am To 5 am'.","","","5AM boundary by design"),
("TC-RD04","Daily Rpt","Empty date — no crash","No data for 2020-01-01","Select 2020-01-01","Date: 2020-01-01","Report shows empty rows or 'No data' message. No error.","","",""),
("TC-RD05","Daily Rpt","Multi-area report","Multiple areas configured","Select Plant001 + AreaA + AreaB","N/A","Tags from both areas shown in single report.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"▶  CALCULATION VERIFICATION — DAILY AVG / MAX / MIN  (Manual Cross-Check Required)", colour=C_PURPLE)
data=[
("TC-RD06","Daily-Calc","Hourly AVG matches DB","Historian data for 2026-05-19","STEP 1 — Query raw DB:\nSELECT AVG(value_num) AS db_avg\nFROM historian_raw.historian_timeseries\nWHERE tag_id = 'Random.Real4'\n  AND time >= '2026-05-19 06:00:00+05:30'\n  AND time  < '2026-05-19 07:00:00+05:30'\n\nSTEP 2 — Note report value for '6am To 7am' column for tag Random.Real4\n\nSTEP 3 — Compare","DB AVG: <calculated value>\nReport value: <shown in UI>","Report value = DB AVG rounded to 2 decimal places.\n|report_val - db_avg| ≤ 0.01","","","PURPLE = manual calc step"),
("TC-RD07","Daily-Calc","Hourly MAX matches DB","Same as above","STEP 1:\nSELECT MAX(value_num)\nFROM historian_raw.historian_timeseries\nWHERE tag_id='Random.Real4'\n  AND time>='2026-05-19 06:00:00+05:30'\n  AND time < '2026-05-19 07:00:00+05:30'\n\nSTEP 2 — Compare to row 'max' field in API response","DB MAX: <calculated>\nReport max: <from API>","Report max = DB MAX (rounded 2dp)","","",""),
("TC-RD08","Daily-Calc","Hourly MIN matches DB","Same as above","STEP 1:\nSELECT MIN(value_num)\nFROM historian_raw.historian_timeseries\nWHERE tag_id='Random.Real4'\n  AND time>='2026-05-19 06:00:00+05:30'\n  AND time < '2026-05-19 07:00:00+05:30'\n\nSTEP 2 — Compare to row 'min' field","DB MIN: <calculated>\nReport min: <from API>","Report min = DB MIN (rounded 2dp)","","",""),
("TC-RD09","Daily-Calc","Day AVG = mean of all 24 hourly AVGs","Data for full day","STEP 1: Get all 24 hourly avgs from API response ('hourly' array)\nSTEP 2: Calculate manually: sum(non-null)/count(non-null)\nSTEP 3: Compare to 'avg' field in API response","Row avg from API: <value>\nManual calc: sum/count=<value>","API 'avg' = round(sum(hourly_avgs_with_data)/count, 2)","","","Formula: row_avg = round(sum(all_avg)/len(all_avg), 2)"),
("TC-RD10","Daily-Calc","Day MAX = max across all hourly maxes","Data for full day","STEP 1: From API row, collect all 24 hourly max_val values\nSTEP 2: Take max of those\nSTEP 3: Compare to row 'max'","Row max: <from API>\nManual: max([h1_max,...,h24_max])","API 'max' = max of all hourly maxes.","","",""),
("TC-RD11","Daily-Calc","Day MIN = min across all hourly mins","Data for full day","STEP 1: From API row, collect all 24 hourly min_val values\nSTEP 2: Take min of those\nSTEP 3: Compare to row 'min'","Row min: <from API>\nManual: min([h1_min,...,h24_min])","API 'min' = min of all hourly mins.","","",""),
("TC-RD12","Daily-Calc","Null hour = not counted in avg","Tag missing data for some hours","Check a tag with missing hours\nVerify: avg = sum / (count of NON-null hours only)","Tag with 20 non-null hours out of 24","Avg uses denominator = 20, not 24","","","Boundary: only non-null hours counted"),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0,hi="calc")

row=sec(ws,row,"EXPORT")
data=[
("TC-RD13","Daily Rpt","Export Excel download","Report loaded","Click Export Excel","N/A",".xlsx file downloads. Opens in Excel. Has hourly columns + avg/max/min.","","",""),
("TC-RD14","API","GET /api/reports/daily/export","Flask running","GET export endpoint","GET /api/reports/daily/export?date=2026-05-19&plant=Plant001&area=AreaA","HTTP 200. Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet","","",""),
("TC-RD15","Daily-Calc","Exported Excel avg matches UI","Report exported","1. Open downloaded Excel\n2. Pick any tag row\n3. Manually =AVERAGE() of the 24 hourly cells in Excel\n4. Compare to Avg column in same row","Excel formula result: <value>\nAvg column: <value>","Excel manual AVERAGE of non-empty cells = Avg column value (±0.01)","","","END-TO-END calc check"),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

# ═══════════════════════════════════════════════════════════════
# 6 — SHIFT REPORT
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("6_Report_Shift")
row=setup_sheet(ws,"Module 6 — Shift Report (Morning 06-14 / Evening 14-22 / Night 22-06)")
row=sec(ws,row,"UI GENERATION")
data=[
("TC-RS01","Shift Rpt","Shift page loads","Flask running","Navigate to Reports > Shift","N/A","Shift selector (Morning/Evening/Night), date picker visible.","","",""),
("TC-RS02","Shift Rpt","Morning shift report","Data for 06:00–14:00","Select 2026-05-19, Morning","Date: 2026-05-19\nShift: Morning (06:00–14:00 IST)","Report shows 8 hourly columns (06→13). Tags from template.","","",""),
("TC-RS03","Shift Rpt","Evening shift report","Data for 14:00–22:00","Select 2026-05-19, Evening","Date: 2026-05-19\nShift: Evening (14:00–22:00 IST)","Report shows 8 hourly columns (14→21).","","",""),
("TC-RS04","Shift Rpt","Night shift report","Data spanning midnight","Select 2026-05-19, Night","Date: 2026-05-19\nShift: Night (22:00 2026-05-19 → 06:00 2026-05-20 IST)","Report covers 22:00–05:59 crossing midnight correctly.","","","Cross-midnight boundary"),
("TC-RS05","API","GET /api/reports/shift","Flask running","GET with params","GET /api/reports/shift?date=2026-05-19&shift=morning&plant=Plant001&area=AreaA","HTTP 200. JSON report data for that shift window.","","",""),
("TC-RS06","API","GET /api/reports/shifts (list)","Flask running","GET shift definitions","GET /api/reports/shifts","HTTP 200. [{id, name, start_time, end_time},...] — at least 3 shifts.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"▶  SHIFT CALCULATION VERIFICATION (Manual Cross-Check)", colour=C_PURPLE)
data=[
("TC-RS07","Shift-Calc","Shift AVG matches DB for window","Data for morning shift","STEP 1:\nSELECT AVG(value_num)\nFROM historian_raw.historian_timeseries\nWHERE tag_id='Random.Real4'\n  AND time>='2026-05-19 06:00:00+05:30'\n  AND time < '2026-05-19 14:00:00+05:30'\n\nSTEP 2: Compare to shift report 'avg' for that tag","DB window avg: <calculated>\nShift report avg: <from UI>","Match within 0.01","","",""),
("TC-RS08","Shift-Calc","Night shift crosses midnight correctly","Night shift data","STEP 1:\nSELECT COUNT(*)\nFROM historian_raw.historian_timeseries\nWHERE tag_id='Random.Real4'\n  AND time>='2026-05-19 22:00:00+05:30'\n  AND time < '2026-05-20 06:00:00+05:30'\n\nSTEP 2: Verify report includes data from both sides of midnight","Row count: should span 2 calendar days","Night shift report includes both 22:00–23:59 AND 00:00–05:59 data.","","","Cross-midnight calc"),
("TC-RS09","Shift-Calc","Shift MAX / MIN correct","Morning shift data","STEP 1:\nSELECT MAX(value_num), MIN(value_num)\nFROM historian_raw.historian_timeseries\nWHERE tag_id='Random.Real4'\n  AND time>='2026-05-19 06:00:00+05:30'\n  AND time < '2026-05-19 14:00:00+05:30'\n\nSTEP 2: Compare to shift report max/min","DB max: <x>, min: <y>\nReport max: <x>, min: <y>","Values match.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0,hi="calc")

row=sec(ws,row,"EXPORT")
data=[
("TC-RS10","Shift Rpt","Export shift Excel","Report loaded","Click Export","N/A","Excel downloads. Has shift window columns + avg/max/min.","","",""),
("TC-RS11","Shift-Calc","Excel avg matches DB","Downloaded Excel","Manually =AVERAGE() shift columns in Excel\nCompare to Avg column","Same tag row","Match ±0.01","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

# ═══════════════════════════════════════════════════════════════
# 7 — MONTHLY REPORT
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("7_Report_Monthly")
row=setup_sheet(ws,"Module 7 — Monthly Report (Daily Rollup Averages)")
row=sec(ws,row,"UI GENERATION")
data=[
("TC-RM01","Monthly Rpt","Monthly page loads","Flask running","Navigate to Reports > Monthly","N/A","Month/Year picker + plant/area dropdowns visible.","","",""),
("TC-RM02","Monthly Rpt","Generate May 2026","Data for May","Select May 2026, Plant001, AreaA","Month: 2026-05\nPlant: Plant001\nArea: AreaA","Table: columns = days 1–31 (or max days in month). Rows = tags.","","",""),
("TC-RM03","Monthly Rpt","No data month","Empty period","Select 2020-01","Month: 2020-01","'No data available' or empty rows. No crash.","","",""),
("TC-RM04","API","GET /api/reports/monthly","Flask running","GET with params","GET /api/reports/monthly?year=2026&month=5&plant=Plant001&area=AreaA","HTTP 200. JSON with daily avg per tag per day.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"▶  MONTHLY CALCULATION VERIFICATION (Manual Cross-Check)", colour=C_PURPLE)
data=[
("TC-RM05","Monthly-Calc","Daily avg in monthly matches daily report","Both reports available","STEP 1: Generate Daily Report for 2026-05-19\nNote avg for tag 'Random.Real4' = X\n\nSTEP 2: Generate Monthly Report for May 2026\nNote value for column '19' for same tag = Y\n\nSTEP 3: X should equal Y","Daily avg (X): <from daily report>\nMonthly day 19 (Y): <from monthly report>","X = Y (same daily average used in both)","","","KEY cross-check: daily and monthly must agree"),
("TC-RM06","Monthly-Calc","Monthly AVG = mean of all daily avgs","May 2026 data","STEP 1: From Monthly report, collect daily column values for one tag\nSTEP 2: =AVERAGE() in Excel over non-blank day cells\nSTEP 3: Compare to 'Monthly Avg' column (if shown)","Manual avg: <calculated>\nReport avg: <shown>","Monthly avg = mean of daily averages (non-null days only)","","",""),
("TC-RM07","Monthly-Calc","Day 19 value in monthly = DB daily avg","DB accessible","STEP 1:\nSELECT AVG(avg_val)\nFROM historian_raw.v_daily_hourly_agg\nWHERE tag_id='Random.Real4' AND local_date='2026-05-19'\n\nSTEP 2: Compare to monthly report column 19 for same tag","DB value: <calculated>\nMonthly col 19: <from report>","Match within 0.01","","","Goes through v_daily_hourly_agg view"),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0,hi="calc")

row=sec(ws,row,"EXPORT")
data=[
("TC-RM08","Monthly Rpt","Export monthly Excel","Report loaded","Click Export","N/A","Excel downloads. Day columns + avg/max/min per row.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

# ═══════════════════════════════════════════════════════════════
# 8 — DATA MATCHING (NEW)
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("8_DataMatching_CalcVerify")
row=setup_sheet(ws,"Module 8 — DATA MATCHING & Manual Calculation Verification")
row=sec(ws,row,"PURPOSE: Verify that every value shown in the UI can be traced back to raw DB rows",colour=C_PURPLE)

row=sec(ws,row,"A — HOURLY AGGREGATE VIEW (v_daily_hourly_agg) vs RAW TABLE", colour=C_PURPLE)
data=[
("TC-DM01","DataMatch","Hourly agg AVG matches raw rows","DB running","STEP 1 — Raw query:\nSELECT AVG(value_num) AS raw_avg, COUNT(*) AS cnt\nFROM historian_raw.historian_timeseries\nWHERE tag_id = 'Random.Real4'\n  AND time >= '2026-05-19 08:00:00+05:30'\n  AND time  < '2026-05-19 09:00:00+05:30'\n\nSTEP 2 — View query:\nSELECT avg_val, min_val, max_val\nFROM historian_raw.v_daily_hourly_agg\nWHERE tag_id='Random.Real4'\n  AND local_date='2026-05-19'\n  AND local_hour=8\n\nSTEP 3 — Compare","Raw AVG: <fill in>\nView avg_val: <fill in>\nDifference: <fill in>","raw_avg = avg_val (within float precision ±0.0001)\nraw MAX = max_val\nraw MIN = min_val","","","Start here for all calculation checks"),
("TC-DM02","DataMatch","Row count in hourly bucket","DB running","SELECT COUNT(*) FROM historian_raw.historian_timeseries\nWHERE tag_id='Random.Real4'\n  AND time>='2026-05-19 08:00:00+05:30'\n  AND time < '2026-05-19 09:00:00+05:30'","Expected: ~60 rows (1 per minute) or ~3600 rows (1 per second)","Count matches expected polling interval. With 1000ms polling = ~60 rows/hour.","","","Validates data ingestion rate"),
("TC-DM03","DataMatch","TimescaleDB hourly agg matches view","DB running","STEP 1:\nSELECT avg_value, min_value, max_value, sample_count\nFROM historian_raw.ts_hourly_agg\nWHERE tag_id='Random.Real4'\n  AND bucket = '2026-05-19 08:00:00+05:30'\n\nSTEP 2 — Compare to v_daily_hourly_agg same hour","ts_hourly_agg avg: <fill>\nv_daily view avg: <fill>","Both views agree within ±0.01 (both derive from same raw table)","","","TimescaleDB native agg vs SQL view"),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0,hi="calc")

row=sec(ws,row,"B — DAILY REPORT API RESPONSE vs DB", colour=C_PURPLE)
data=[
("TC-DM04","DataMatch","API hourly column matches DB avg","DB + Flask running","STEP 1: Call API:\nGET /api/reports/daily?date=2026-05-19&plant=Plant001&area=AreaA\n\nSTEP 2: Find tag 'Random.Real4' in JSON\nNote hourly[2] (index 2 = 7am–8am, 3rd column)\n\nSTEP 3: Query DB:\nSELECT AVG(value_num)\nFROM historian_raw.historian_timeseries\nWHERE tag_id='Random.Real4'\n  AND time>='2026-05-19 07:00:00+05:30'\n  AND time < '2026-05-19 08:00:00+05:30'\n\nSTEP 4: Compare","API hourly[2]: <fill>\nDB AVG: <fill>","Match within 0.01","","","Use ordered hours: index 0=5am, 1=6am, 2=7am..."),
("TC-DM05","DataMatch","Row 'avg' = mean of non-null hourly[]","API response","STEP 1: Get hourly[] array for one tag from API\nSTEP 2: Filter out null values\nSTEP 3: Python: round(sum(non_null)/len(non_null), 2)\nSTEP 4: Compare to 'avg' field in same row","hourly[] non-null values: [12.3, 13.1, ...]\nManual avg: <calculate>\nAPI avg: <from response>","API avg = manual round(avg, 2)","","","Core formula check"),
("TC-DM06","DataMatch","Row 'max' = max of all hourly max_vals","API response","STEP 1: For same tag, from v_daily_hourly_agg:\nSELECT MAX(max_val)\nFROM historian_raw.v_daily_hourly_agg\nWHERE tag_id='Random.Real4' AND local_date='2026-05-19'\n\nSTEP 2: Compare to 'max' in API response","DB max of maxes: <fill>\nAPI row max: <fill>","Match","","",""),
("TC-DM07","DataMatch","Row 'min' = min of all hourly min_vals","API response","STEP 1:\nSELECT MIN(min_val)\nFROM historian_raw.v_daily_hourly_agg\nWHERE tag_id='Random.Real4' AND local_date='2026-05-19'\n\nSTEP 2: Compare to 'min' in API response","DB min of mins: <fill>\nAPI row min: <fill>","Match","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0,hi="calc")

row=sec(ws,row,"C — MONTHLY vs DAILY CROSS-CHECK", colour=C_PURPLE)
data=[
("TC-DM08","DataMatch","Monthly day cell = daily report avg","Both reports generated","STEP 1: Daily report 2026-05-15, tag Random.Real4 → note 'avg' = A\nSTEP 2: Monthly report May 2026, column '15', same tag → note value = B\nSTEP 3: A == B?","Daily avg (A): <fill>\nMonthly col 15 (B): <fill>","A = B","","","Critical cross-report consistency"),
("TC-DM09","DataMatch","All 3 reports pull from same source table","DB running","Run same raw query 3 ways:\n1. Direct: SELECT AVG(value_num)...\n2. Via v_daily_hourly_agg\n3. Via ts_hourly_agg\nCompare all 3","All 3 values: <fill each>","All 3 match within ±0.01. Confirms no separate calculation paths.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0,hi="calc")

row=sec(ws,row,"D — LIVE VALUE vs HISTORIAN DB", colour=C_PURPLE)
data=[
("TC-DM10","DataMatch","Live API value appears in DB within 2s","OPC connected, tag in tag_master","STEP 1: Read current value from GET /api/opc/values for Random.Real4 → note value V and timestamp T\nSTEP 2: Wait 3 seconds\nSTEP 3: Query:\nSELECT value_num, time FROM historian_raw.historian_timeseries\nWHERE tag_id='Random.Real4' ORDER BY time DESC LIMIT 3\nSTEP 4: Check if V appears near timestamp T","Live value V: <fill>\nTimestamp T: <fill>\nFound in DB: YES/NO","V found in DB within ±2s of T. Confirms ingest pipeline working.","","","End-to-end live → DB check"),
("TC-DM11","DataMatch","Deadband prevents duplicate — manual check","Tag with deadband=0.5","STEP 1: Note last 10 DB values for tag:\nSELECT value_num, time FROM historian_raw.historian_timeseries\nWHERE tag_id='Random.Real4' ORDER BY time DESC LIMIT 10\nSTEP 2: Check that consecutive values differ by > 0.5\n(or first sample always written)","10 consecutive values: <fill column>","No two consecutive rows have |delta| < deadband_value (except first sample or spike)","","","RateController verification"),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0,hi="calc")

# ═══════════════════════════════════════════════════════════════
# 9 — DB LOGS VERIFICATION (NEW)
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("9_DB_Logs_Verification")
row=setup_sheet(ws,"Module 9 — Database Log & Write Verification")
row=sec(ws,row,"A — HISTORIAN TIMESERIES — WRITE VERIFICATION")
data=[
("TC-DB01","DB-Log","New rows written every ~1s","OPC connected, tag enabled in tag_master","Note current count:\nSELECT count(*) FROM historian_raw.historian_timeseries WHERE tag_id='Random.Real4' AND time > now()-interval '10s'\n\nWait 10 seconds. Run again.","Count at T=0: <fill>\nCount at T=10s: <fill>","Count increased by ~10 rows (1 per second). Confirms ingest active.","","",""),
("TC-DB02","DB-Log","No writes when OPC disconnected","OPC disconnected","1. Stop OPC server\n2. Wait 30s\n3. Check count","SELECT count(*) WHERE time > now()-interval '30s'","Count = 0 or very low. No fake values written.","","",""),
("TC-DB03","DB-Log","Quality column values valid","DB running","SELECT DISTINCT quality FROM historian_raw.historian_timeseries","N/A","Only valid values: 'G' (Good), 'B' (Bad), 'U' (Uncertain), 'C' (CommError)","","",""),
("TC-DB04","DB-Log","No NULL tag_id rows","DB running","SELECT count(*) FROM historian_raw.historian_timeseries WHERE tag_id IS NULL","N/A","Returns 0. No orphan rows without tag ID.","","",""),
("TC-DB05","DB-Log","Timestamps in expected timezone","DB running","SELECT time, time AT TIME ZONE 'Asia/Kolkata' AS ist_time\nFROM historian_raw.historian_timeseries\nORDER BY time DESC LIMIT 5","N/A","ist_time shows current IST time. Confirms timezone-aware storage.","","",""),
("TC-DB06","DB-Log","No future timestamps","DB running","SELECT count(*) FROM historian_raw.historian_timeseries WHERE time > now() + interval '1 minute'","N/A","Returns 0. No corrupt future-dated rows.","","",""),
("TC-DB07","DB-Log","No duplicate (tag_id, time) pairs","DB running","SELECT tag_id, time, count(*) FROM historian_raw.historian_timeseries GROUP BY tag_id, time HAVING count(*) > 1 LIMIT 5","N/A","Returns 0 rows. No duplicates.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"B — HOURLY AGGREGATE VIEW (v_daily_hourly_agg)")
data=[
("TC-DB08","DB-Log","Agg view has data for yesterday","DB running","SELECT count(*) FROM historian_raw.v_daily_hourly_agg WHERE local_date = current_date - 1","N/A","Returns > 0. Yesterday's aggregates exist.","","",""),
("TC-DB09","DB-Log","All 24 hours populated for a full day","DB running","SELECT local_hour, count(DISTINCT tag_id) FROM historian_raw.v_daily_hourly_agg\nWHERE local_date='2026-05-19'\nGROUP BY local_hour ORDER BY local_hour","N/A","24 rows (hours 0–23), each with > 0 tags.","","",""),
("TC-DB10","DB-Log","avg_val not null for populated hours","DB running","SELECT count(*) FROM historian_raw.v_daily_hourly_agg\nWHERE local_date='2026-05-19' AND avg_val IS NULL","N/A","Returns 0 (or very low — only if a tag had no data that hour)","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"C — TimescaleDB CONTINUOUS AGGREGATE (ts_hourly_agg)")
data=[
("TC-DB11","DB-Log","ts_hourly_agg refreshes after new data","DB running","1. Insert test row\n2. Wait up to 1h (or manually refresh)\n3. Query ts_hourly_agg for that bucket","CALL refresh_continuous_aggregate('historian_raw.ts_hourly_agg', now()-interval '2h', now())","New bucket row appears in ts_hourly_agg.","","","Auto-refreshes every 1h"),
("TC-DB12","DB-Log","sample_count in ts_hourly_agg reasonable","DB running","SELECT tag_id, bucket, sample_count FROM historian_raw.ts_hourly_agg WHERE tag_id='Random.Real4' ORDER BY bucket DESC LIMIT 5","N/A","sample_count ≈ 60 (for 1 sample/min) or 3600 (1/sec). Consistent per tag.","","",""),
("TC-DB13","DB-Log","ts_hourly_agg avg matches v_daily_hourly_agg","DB running","SELECT a.avg_value AS ts_avg, b.avg_val AS view_avg\nFROM historian_raw.ts_hourly_agg a\nJOIN historian_raw.v_daily_hourly_agg b\n  ON a.tag_id=b.tag_id AND EXTRACT(HOUR FROM a.bucket AT TIME ZONE 'Asia/Kolkata')=b.local_hour AND DATE(a.bucket AT TIME ZONE 'Asia/Kolkata')=b.local_date\nWHERE a.tag_id='Random.Real4'\nLIMIT 5","N/A","ts_avg = view_avg within ±0.001 for each row","","","Critical: both aggs must agree"),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"D — ALARM DB LOGS")
data=[
("TC-DB14","DB-Log","Active alarms table has correct schema","DB running","SELECT column_name, data_type FROM information_schema.columns WHERE table_name='alarm_active'","N/A","Columns include: id, tag_id, alarm_time, severity, ack_status, operator","","",""),
("TC-DB15","DB-Log","ACK creates audit trail row","Alarm ACKed via UI","After ACK, query:\nSELECT * FROM historian_raw.alarm_audit_trail WHERE alarm_id=<N> ORDER BY event_time DESC LIMIT 1","Alarm ID: <from active list>","Row with action='ACKNOWLEDGED', operator='Mustafa', timestamp=now","","",""),
("TC-DB16","DB-Log","Suppressed alarm in alarm_active with status","Alarm suppressed","SELECT status, suppressed_until FROM historian_raw.alarm_active WHERE id=<N>","N/A","status='suppressed', suppressed_until = configured end time","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"E — AUDIT TRAIL DB LOGS")
data=[
("TC-DB17","DB-Log","Login creates user_sessions row","Logged in","SELECT * FROM historian_meta.user_sessions ORDER BY created_at DESC LIMIT 3","N/A","Most recent row shows user_id=Mustafa, created_at=now, is_active=true","","",""),
("TC-DB18","DB-Log","Logout closes session in DB","Logged out","SELECT is_active, ended_at FROM historian_meta.user_sessions ORDER BY created_at DESC LIMIT 1","N/A","is_active=false, ended_at=logout timestamp","","",""),
("TC-DB19","DB-Log","Session activity log has entries","Logged in, navigated","SELECT * FROM historian_meta.session_activity_log ORDER BY logged_at DESC LIMIT 5","N/A","Rows with user, action, endpoint, timestamp","","",""),
("TC-DB20","DB-Log","Report generation logged","Report generated via UI","SELECT * FROM historian_meta.report_gen_log ORDER BY generated_at DESC LIMIT 3","N/A","Row shows: report_type, plant, area, date, generated_by=Mustafa, generated_at=now","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"F — TAG MASTER & MAPPING LOGS")
data=[
("TC-DB21","DB-Log","tag_master enabled flags checked","DB running","SELECT tag_id, enabled, include_in_report FROM historian_meta.tag_master WHERE enabled=TRUE LIMIT 10","N/A","All rows: enabled=true. Shows tags actively being logged.","","",""),
("TC-DB22","DB-Log","Disabled tag has no recent DB writes","DB running, disabled tag","SELECT count(*) FROM historian_raw.historian_timeseries WHERE tag_id='<disabled_tag>' AND time > now()-interval '5m'","tag_id: a tag with enabled=false","Returns 0. Confirms rate controller respects enabled flag.","","",""),
("TC-DB23","DB-Log","include_in_report tags appear in daily report","DB running","STEP 1: SELECT tag_id FROM historian_meta.tag_master WHERE include_in_report=TRUE LIMIT 5\nSTEP 2: Generate daily report\nSTEP 3: Confirm those tag_ids appear in report rows","Expected tags: <list from DB>","Each tag_id with include_in_report=TRUE appears in the report.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

# ═══════════════════════════════════════════════════════════════
# 10 — ADMIN
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("10_Admin")
row=setup_sheet(ws,"Module 10 — Admin Panel (Users / Roles / Permissions / Alerts)")
row=sec(ws,row,"USER MANAGEMENT")
data=[
("TC-AD01","Admin","Admin panel accessible — Admin role","Logged in as Mustafa (Admin)","Navigate to /admin","N/A","Admin panel loads. Tabs: Users, Roles, Permissions, Alerts.","","",""),
("TC-AD02","Admin","Non-admin role blocked","Logged in as non-admin","Navigate to /admin","N/A","Redirect/403. 'Access Denied' shown.","","",""),
("TC-AD03","Admin","Users tab — list all users","Admin logged in","Click Users tab","N/A","Table: username, email, role, status, MFA enabled, last login.","","",""),
("TC-AD04","Admin","Create new user","Admin logged in","1. Click Add User\n2. Fill all fields\n3. Submit","Username: testop01\nEmail: op@test.com\nRole: Operator\nPassword: Test@12345","User appears in list. Login with those creds works.","","",""),
("TC-AD05","Admin","Disable user account","testop01 exists","1. Click Disable on testop01","N/A","testop01 status = disabled. Login attempt: 'Account disabled'.","","",""),
("TC-AD06","Admin","Re-enable user","testop01 disabled","Click Enable","N/A","User can login again.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

row=sec(ws,row,"ROLES & PERMISSIONS")
data=[
("TC-AD07","Admin","Roles tab lists all roles","Admin logged in","Click Roles tab","N/A","Table: role name, description, permission count. Minimum: Admin, Operator, Engineer, Viewer.","","",""),
("TC-AD08","Admin","Assign role to user","User + role exist","1. Edit testop01\n2. Change role to Engineer","New role: Engineer","Role updated. testop01 sees Engineer-level menus after re-login.","","",""),
("TC-AD09","Admin","Tag permission — restrict for role","Permissions tab open","1. Remove tag Random.Real4 from Operator role\n2. Login as Operator user","Tag: Random.Real4\nRole: Operator","Operator dashboard no longer shows Random.Real4.","","",""),
("TC-AD10","Admin","System alerts tab","Admin logged in","Click Alerts tab","N/A","System alert configuration visible. No JS error.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

# ═══════════════════════════════════════════════════════════════
# 11 — ASSET BROWSER
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("11_AssetBrowser")
row=setup_sheet(ws,"Module 11 — Asset Browser & Equipment Hierarchy")
row=sec(ws,row,"HIERARCHY TREE")
data=[
("TC-AS01","Assets","Asset browser loads","Hierarchy seeded in DB","Navigate to Asset Browser page","N/A","Tree shows: Plant > Area > Equipment > Sub-equipment nodes.","","",""),
("TC-AS02","Assets","Expand plant node","Hierarchy data exists","Click Plant001 node","N/A","Expands to show Areas under Plant001.","","",""),
("TC-AS03","Assets","Expand area node","Areas have equipment","Click AreaA node","N/A","Equipment list shown (Turbine, Boiler, etc.)","","",""),
("TC-AS04","Assets","Click equipment — live tags panel","Equipment has tags","Click Turbine001","N/A","Right panel shows live tag values for Turbine001. Values updating.","","",""),
("TC-AS05","Assets","Search equipment","Many assets","Type in search box: 'Turbine'","Search: Turbine","Tree filters to matching nodes only.","","",""),
("TC-AS06","API","GET /api/equipment/hierarchy","Flask running","GET request","GET http://localhost:6001/api/equipment/hierarchy\nHeader: Bearer <token>","HTTP 200. Nested JSON {plant:[{area:[{equipment:[...]}]}]}","","",""),
("TC-AS07","API","GET /api/asset — asset list","Flask running","GET asset list","GET /api/asset","HTTP 200. Array of asset objects.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

# ═══════════════════════════════════════════════════════════════
# 12 — AUDIT TRAIL
# ═══════════════════════════════════════════════════════════════
ws=wb.create_sheet("12_AuditTrail")
row=setup_sheet(ws,"Module 12 — Audit Trail")
row=sec(ws,row,"AUDIT LOG UI & API")
data=[
("TC-AU01","Audit","Audit tab/page loads","Actions logged","Navigate to Audit section","N/A","Table: timestamp, user, action, target object, IP address.","","",""),
("TC-AU02","Audit","Login creates audit entry","Audit active","1. Login as Mustafa\n2. Check audit log","N/A","Entry: action='LOGIN', user='Mustafa', timestamp≈now.","","",""),
("TC-AU03","Audit","Alarm ACK creates audit","Alarm ACKed","1. ACK alarm\n2. Check audit log","Alarm ID: any","Entry: action='ALARM_ACK', operator='Mustafa', alarm_id=N.","","",""),
("TC-AU04","Audit","Report generation logged","Report generated","1. Generate daily report\n2. Check report_gen_log","N/A","Entry in historian_meta.report_gen_log with generated_by=Mustafa.","","",""),
("TC-AU05","Audit","Filter by username","Entries exist","1. Filter: user=Mustafa","Filter: user=Mustafa","Only Mustafa's entries shown.","","",""),
("TC-AU06","Audit","Filter by date range","Entries exist","Filter: today only","Start: today 00:00\nEnd: today 23:59","Only today's entries shown.","","",""),
("TC-AU07","API","GET /api/audit","Flask running","GET with filter","GET /api/audit?user=Mustafa&limit=50","HTTP 200. Array of up to 50 audit events for Mustafa.","","",""),
("TC-AU08","API","GET /api/audit/operator/<name>/stats","Flask running","GET operator stats","GET /api/alarms/audit/operator/Mustafa/stats","HTTP 200. Count of actions by type for Mustafa.","","",""),
]
for i,d in enumerate(data): row=row_write(ws,row,d,alt=i%2==0)

# ─── SAVE ─────────────────────────────────────────────────────
wb.save("TEST_FUNCTIONS_HMI_COMPREHENSIVE.xlsx")
print("\n  Done: TEST_FUNCTIONS_HMI_COMPREHENSIVE.xlsx\n")
print("  Sheets:")
for s in wb.sheetnames:
    print(f"    {s}")
