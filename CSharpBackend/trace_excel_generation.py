"""
COMPREHENSIVE EXCEL GENERATION DEBUG SCRIPT
This script traces the ENTIRE flow from database to Excel file
"""

import psycopg2
import sys
from io import BytesIO
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

print("=" * 100)
print("COMPREHENSIVE EXCEL GENERATION DEBUG - TRACING ENTIRE FLOW")
print("=" * 100)

# Database connection
conn = psycopg2.connect(
    host="localhost",
    port=5432,
    database="Automation_DB",
    user="cereveate",
    password="cereveate@222"
)
cursor = conn.cursor()

# Test parameters
report_date = date(2026, 5, 18)
plants = ["FTP-1"]
areas = ["POTLINE"]

print(f"\n[TEST PARAMS]")
print(f"  Date: {report_date}")
print(f"  Plants: {plants}")
print(f"  Areas: {areas}")

# STEP 1: Simulate build_daily_report query
print(f"\n[STEP 1] Building report data (simulating report_service.py logic)...")

# Template query (same as report_service.py lines 82-109)
template_query = """
SELECT vt.s_no, vt.tag_id, vt.display_label, vt.group_name, vt.parameter_unit,
       vt.plant, vt.area, tm.sub_equipment, tm.description, tm.eng_unit
FROM historian_meta.v_report_template_tags vt
LEFT JOIN historian_meta.tag_master tm ON vt.tag_id = tm.tag_id
WHERE vt.plant = ANY(%s) AND vt.area = ANY(%s)
  AND vt.tag_id IN (
    SELECT DISTINCT tag_id 
    FROM historian_raw.v_daily_hourly_agg
    WHERE local_date = %s
  )
ORDER BY vt.s_no;
"""

cursor.execute(template_query, (plants, areas, report_date))
template_rows = cursor.fetchall()

print(f"  Template query returned: {len(template_rows)} tags")

if len(template_rows) == 0:
    print("  ⚠️ No tags from template, trying fallback query...")
    
    fallback_query = """
    SELECT DISTINCT 
           ROW_NUMBER() OVER (ORDER BY tm.tag_id) as s_no,
           tm.tag_id, tm.tag_name as display_label, tm.equipment as group_name,
           tm.eng_unit as parameter_unit, tm.plant, tm.area,
           tm.sub_equipment, tm.description, tm.eng_unit
    FROM historian_meta.tag_master tm
    WHERE tm.plant = ANY(%s) AND tm.area = ANY(%s)
      AND tm.tag_id IN (
        SELECT DISTINCT tag_id 
        FROM historian_raw.v_daily_hourly_agg
        WHERE local_date = %s
      )
    ORDER BY tm.tag_id;
    """
    
    cursor.execute(fallback_query, (plants, areas, report_date))
    template_rows = cursor.fetchall()
    print(f"  Fallback query returned: {len(template_rows)} tags")

if len(template_rows) == 0:
    print("  ❌ NO DATA FOUND! Check database has data for this date/plant/area")
    cursor.close()
    conn.close()
    sys.exit(1)

# Parse template rows
tags_data = {}
for row in template_rows:
    s_no, tag_id, display_label, group_name, parameter_unit, plant, area, sub_equipment, description, eng_unit = row
    tags_data[tag_id] = {
        's_no': s_no,
        'tag_id': tag_id,
        'display_label': display_label,
        'group_name': group_name,
        'parameter_unit': parameter_unit,
        'plant': plant,
        'area': area,
        'sub_equipment': sub_equipment,
        'description': description,
        'eng_unit': eng_unit
    }

print(f"\n[STEP 1a] Inspecting first tag data BEFORE clean_field:")
first_tag = list(tags_data.values())[0]
print(f"  tag_id: '{first_tag['tag_id']}'")
print(f"  group_name: '{first_tag['group_name']}'")
print(f"  sub_equipment: '{first_tag['sub_equipment']}'")
print(f"  description: '{first_tag['description']}'")
print(f"  eng_unit: '{first_tag['eng_unit']}'")

# Check for 'None' strings
if first_tag['sub_equipment'] == 'None':
    print("  ⚠️ sub_equipment is string 'None'")
if first_tag['description'] == 'None':
    print("  ⚠️ description is string 'None'")
if first_tag['eng_unit'] == 'None':
    print("  ⚠️ eng_unit is string 'None'")

# STEP 2: Get hourly data
print(f"\n[STEP 2] Fetching hourly aggregated data...")

hourly_query = """
SELECT tag_id, local_hour, 
       ROUND(CAST(avg_val AS numeric), 2) as avg_val,
       ROUND(CAST(max_val AS numeric), 2) as max_val,
       ROUND(CAST(min_val AS numeric), 2) as min_val
FROM historian_raw.v_daily_hourly_agg
WHERE local_date = %s AND tag_id = ANY(%s)
ORDER BY tag_id, local_hour;
"""

tag_ids = list(tags_data.keys())
cursor.execute(hourly_query, (report_date, tag_ids))
hourly_rows = cursor.fetchall()

print(f"  Fetched {len(hourly_rows)} hourly data points")

# Organize hourly data by tag
hourly_by_tag = {}
for tag_id in tag_ids:
    hourly_by_tag[tag_id] = {h: None for h in range(24)}

for tag_id, local_hour, avg_val, max_val, min_val in hourly_rows:
    hourly_by_tag[tag_id][local_hour] = (avg_val, max_val, min_val)

# STEP 3: Build report rows (simulating report_service.py logic)
print(f"\n[STEP 3] Building report rows with clean_field helper...")

def clean_field(value):
    """Helper to filter out 'None' string and None/empty values"""
    if value is None or value == '' or value == 'None':
        return ''
    return value

# Build rows in 5AM-5AM order
hour_order = [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 0, 1, 2, 3, 4]

rows_out = []
for tag_id in sorted(tags_data.keys()):
    trow = tags_data[tag_id]
    
    # Build hourly values in 5AM-5AM order
    hourly_values = []
    all_avg, all_max, all_min = [], [], []
    
    for h in hour_order:
        if hourly_by_tag[tag_id][h]:
            avg_val, max_val, min_val = hourly_by_tag[tag_id][h]
            hourly_values.append(avg_val)
            if avg_val is not None:
                all_avg.append(avg_val)
            if max_val is not None:
                all_max.append(max_val)
            if min_val is not None:
                all_min.append(min_val)
        else:
            hourly_values.append(None)
    
    row_avg = round(sum(all_avg) / len(all_avg), 2) if all_avg else None
    row_max = round(max(all_max), 2) if all_max else None
    row_min = round(min(all_min), 2) if all_min else None
    
    rows_out.append({
        "s_no": trow["s_no"],
        "group": clean_field(trow.get("group_name")),
        "sub_equipment": clean_field(trow.get("sub_equipment")),
        "tag_id": tag_id,
        "description": clean_field(trow.get("description")) or clean_field(trow.get("display_label")) or tag_id,
        "eng_unit": clean_field(trow.get("eng_unit")),
        "parameter_unit": clean_field(trow.get("parameter_unit")),
        "display_label": clean_field(trow.get("display_label")) or tag_id,
        "avg": row_avg,
        "max": row_max,
        "min": row_min,
        "hourly": hourly_values,
    })

print(f"  Built {len(rows_out)} report rows")

print(f"\n[STEP 3a] Inspecting first row AFTER clean_field:")
first_row = rows_out[0]
print(f"  tag_id: '{first_row['tag_id']}'")
print(f"  group: '{first_row['group']}'")
print(f"  sub_equipment: '{first_row['sub_equipment']}'")
print(f"  description: '{first_row['description']}'")
print(f"  eng_unit: '{first_row['eng_unit']}'")

blank_fields = []
if not first_row['sub_equipment']:
    blank_fields.append('sub_equipment')
if not first_row['description']:
    blank_fields.append('description')
if not first_row['eng_unit']:
    blank_fields.append('eng_unit')

if blank_fields:
    print(f"  ⚠️ BLANK AFTER clean_field: {', '.join(blank_fields)}")
else:
    print(f"  ✅ All fields populated after clean_field")

# STEP 4: Generate Excel (simulating export_to_excel)
print(f"\n[STEP 4] Generating Excel file...")

wb = Workbook()
ws = wb.active
ws.title = "Daily Report"

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Row 1: Company name
ws.append(["BHARAT ALUMINIUM COMPANY LIMITED ( PLANT- II )\nPOTLINE, FUME TREATMENT PLANT"])
ws.merge_cells("A1:AD1")
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 30

# Row 2: Report title and date
row2_data = [""] * 30
row2_data[2] = "DAILY REPORT  (CONTROL ROOM)"
row2_data[4] = "DATE"
row2_data[5] = report_date.strftime("%d-%B-%y")
ws.append(row2_data)
ws["C2"].font = Font(bold=True, size=12)
ws["E2"].font = Font(bold=True)
ws["F2"].font = Font(bold=True)

# Row 3: Headers
hour_labels = []
for h in hour_order:
    next_h = (h + 1) % 24
    am_pm1 = "AM" if h < 12 else "PM"
    am_pm2 = "AM" if next_h < 12 else "PM"
    h1 = h if h <= 12 else h - 12
    h2 = next_h if next_h <= 12 else next_h - 12
    if h1 == 0:
        h1 = 12
    if h2 == 0:
        h2 = 12
    hour_labels.append(f"{h1} {am_pm1} \n To {h2} {am_pm2}")

headers_row = ["Equipment", "Sub Equipment", "Tag Name", "Tag Description", "Unit"] + hour_labels + ["MIN", "MAX", "AVG"]
ws.append(headers_row)

# Style headers
header_fill = PatternFill("solid", fgColor="4472C4")
summary_fill = PatternFill("solid", fgColor="E7E6E6")
white_bold = Font(color="FFFFFF", bold=True, size=10)
for cell in ws[3]:
    if cell.value:
        cell.fill = header_fill
        cell.font = white_bold
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data rows
for row in rows_out:
    data_row = [
        clean_field(row.get("group")),                      # Equipment
        clean_field(row.get("sub_equipment")),              # Sub Equipment
        row["tag_id"],                                      # Tag Name
        clean_field(row.get("description")) or clean_field(row.get("display_label")) or row["tag_id"],
        clean_field(row.get("eng_unit")),                   # Unit
        *row["hourly"],                                     # Hourly values
        row["min"],                                         # MIN
        row["max"],                                         # MAX
        row["avg"],                                         # AVG
    ]
    ws.append(data_row)

print(f"  Excel worksheet created with {ws.max_row} rows")

# STEP 5: Verify Excel contents
print(f"\n[STEP 5] Verifying Excel cell contents...")

print(f"  Row 4 (first data row) cell values:")
print(f"    A4 (Equipment): '{ws['A4'].value}'")
print(f"    B4 (Sub Equipment): '{ws['B4'].value}'")
print(f"    C4 (Tag Name): '{ws['C4'].value}'")
print(f"    D4 (Tag Description): '{ws['D4'].value}'")
print(f"    E4 (Unit): '{ws['E4'].value}'")

blank_cells = []
if not ws['B4'].value:
    blank_cells.append('B4 (Sub Equipment)')
if not ws['D4'].value:
    blank_cells.append('D4 (Tag Description)')
if not ws['E4'].value:
    blank_cells.append('E4 (Unit)')

if blank_cells:
    print(f"\n  ❌ BLANK CELLS IN EXCEL: {', '.join(blank_cells)}")
    print(f"\n  🔍 ROOT CAUSE ANALYSIS:")
    print(f"     Database values BEFORE clean_field: sub_equipment='{first_tag['sub_equipment']}', description='{first_tag['description']}', eng_unit='{first_tag['eng_unit']}'")
    print(f"     Row dict AFTER clean_field: sub_equipment='{first_row['sub_equipment']}', description='{first_row['description']}', eng_unit='{first_row['eng_unit']}'")
    print(f"     Excel cell values: B4='{ws['B4'].value}', D4='{ws['D4'].value}', E4='{ws['E4'].value}'")
    print(f"\n  💡 DIAGNOSIS:")
    if first_tag['sub_equipment'] == 'None':
        print(f"     - Database has string 'None' in sub_equipment field")
    if first_tag['description'] == 'None':
        print(f"     - Database has string 'None' in description field")
    if first_tag['eng_unit'] == 'None':
        print(f"     - Database has string 'None' in eng_unit field")
    print(f"     - clean_field() should filter these out but cells are still blank")
    print(f"     - This means clean_field() returned empty string which Excel displays as blank")
else:
    print(f"\n  ✅ ALL CELLS POPULATED IN EXCEL!")

# Save Excel file
output_file = f"debug_excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(output_file)
print(f"\n[STEP 6] Excel file saved: {output_file} ({ws.max_row} rows, {ws.max_column} columns)")

cursor.close()
conn.close()

print("\n" + "=" * 100)
print("DEBUG COMPLETE")
print(f"Open {output_file} to manually verify the data")
print("=" * 100)
